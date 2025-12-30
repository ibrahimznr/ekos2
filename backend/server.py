from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
from passlib.context import CryptContext
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import base64
import zipfile
import shutil
import tempfile
from constants import SEHIRLER, KATEGORI_ALT_KATEGORI

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT & Password
SECRET_KEY = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440  # 24 hours
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Create the main app
app = FastAPI(title="EKOS - Ekipman Kontrol Otomasyon Sistemi")
api_router = APIRouter(prefix="/api")

# Upload directory
UPLOAD_DIR = Path("/app/uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: EmailStr
    password: str
    role: str = "viewer"  # admin, inspector, viewer
    firma_adi: Optional[str] = None  # Firma-based access control
    email_verified: bool = False
    verification_code: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    username: str
    email: EmailStr
    password: str
    password_confirm: str
    firma_adi: str  # Required for registration
    role: str = "viewer"

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    username: str
    email: str
    role: str
    firma_adi: Optional[str] = None
    email_verified: bool
    created_at: datetime

class VerifyEmail(BaseModel):
    email: EmailStr
    code: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse

class Kategori(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    isim: str
    alt_kategoriler: List[str] = []
    aciklama: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class KategoriCreate(BaseModel):
    isim: str
    alt_kategoriler: List[str] = []
    aciklama: Optional[str] = None

class Proje(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    proje_adi: str
    proje_kodu: Optional[str] = None
    lokasyon: Optional[str] = None
    baslangic_tarihi: Optional[str] = None
    bitis_tarihi: Optional[str] = None
    durum: str = "Aktif"
    aciklama: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProjeCreate(BaseModel):
    proje_adi: str
    proje_kodu: str
    lokasyon: Optional[str] = None
    baslangic_tarihi: Optional[str] = None
    bitis_tarihi: Optional[str] = None
    durum: str = "Aktif"
    aciklama: Optional[str] = None

class Rapor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    rapor_no: str
    proje_id: str
    proje_adi: str
    sehir: str
    sehir_kodu: str
    ekipman_adi: str
    kategori: str
    alt_kategori: Optional[str] = None
    firma: str
    lokasyon: Optional[str] = None
    marka_model: Optional[str] = None
    seri_no: Optional[str] = None
    periyot: Optional[str] = None  # 3/6/12 Aylık
    gecerlilik_tarihi: Optional[str] = None
    aciklama: Optional[str] = None
    uygunluk: Optional[str] = None  # Uygun/Uygun Değil
    durum: str = "Aktif"  # Aktif/Pasif
    created_by: str
    created_by_username: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RaporCreate(BaseModel):
    proje_id: str
    sehir: str
    ekipman_adi: str
    kategori: str
    alt_kategori: Optional[str] = None
    firma: str
    lokasyon: Optional[str] = None
    marka_model: Optional[str] = None
    seri_no: Optional[str] = None
    periyot: Optional[str] = None
    gecerlilik_tarihi: Optional[str] = None
    aciklama: Optional[str] = None
    uygunluk: Optional[str] = None

# İskele Bileşeni Model
class IskeleBileseni(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    proje_id: str
    proje_adi: str
    bileşen_adi: str
    malzeme_kodu: str  # Seri no'ya karşılık geliyor
    bileşen_adedi: int
    firma_adi: str
    iskele_periyodu: str = "6 Aylık"  # Sabit değer
    gecerlilik_tarihi: Optional[str] = None
    uygunluk: str = "Uygun"  # Uygun/Uygun Değil
    aciklama: Optional[str] = None
    gorseller: List[str] = Field(default_factory=list)  # Max 3 görsel URL
    created_by: str
    created_by_username: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class IskeleBileseniCreate(BaseModel):
    proje_id: str
    bileşen_adi: str
    malzeme_kodu: str
    bileşen_adedi: int
    firma_adi: str
    gecerlilik_tarihi: Optional[str] = None
    uygunluk: str = "Uygun"
    aciklama: Optional[str] = None
    gorseller: List[str] = Field(default_factory=list)

class RaporUpdate(BaseModel):
    proje_id: Optional[str] = None
    sehir: Optional[str] = None
    ekipman_adi: Optional[str] = None
    kategori: Optional[str] = None
    alt_kategori: Optional[str] = None
    firma: Optional[str] = None
    lokasyon: Optional[str] = None
    marka_model: Optional[str] = None
    seri_no: Optional[str] = None
    periyot: Optional[str] = None
    gecerlilik_tarihi: Optional[str] = None
    aciklama: Optional[str] = None
    uygunluk: Optional[str] = None

class MedyaDosyasi(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    rapor_id: str
    dosya_adi: str
    dosya_yolu: str
    dosya_tipi: str
    dosya_boyutu: int
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Helper Functions
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Geçersiz kimlik")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token süresi dolmuş")
    except jwt.JWTError:
        raise HTTPException(status_code=401, detail="Geçersiz token")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user is None:
        raise HTTPException(status_code=401, detail="Kullanıcı bulunamadı")
    
    # Fallback for old users without username
    if "username" not in user:
        user["username"] = user["email"].split("@")[0]
    
    return user

def get_sehir_kodu(sehir_ismi: str) -> str:
    """Get city code from city name"""
    for sehir in SEHIRLER:
        if sehir['isim'] == sehir_ismi:
            return sehir['kod']
    return "XXX"  # Default if not found

async def generate_rapor_no(sehir: str):
    """
    Generate report number in format: PKYYYY-SEHIRKODU###
    Example: PK2025-ANK025 (for Ankara)
    """
    now = datetime.now(timezone.utc)
    year = now.strftime("%Y")
    
    # Get city code from city name
    sehir_kodu = get_sehir_kodu(sehir)
    prefix = f"PK{year}-{sehir_kodu}"
    
    # Get the last report for this city and year
    last_report = await db.raporlar.find_one(
        {"rapor_no": {"$regex": f"^{prefix}"}},
        sort=[("created_at", -1)]
    )
    
    if last_report:
        last_no_str = last_report["rapor_no"].split(sehir_kodu)[-1]
        last_no = int(last_no_str)
        new_no = last_no + 1
    else:
        new_no = 1
    
    return f"{prefix}{str(new_no).zfill(3)}"

# Initialize default admin
@app.on_event("startup")
async def startup_db():
    # Create indexes for better performance
    try:
        # Users collection indexes
        await db.users.create_index("email", unique=True)
        await db.users.create_index("username")
        
        # Raporlar collection indexes
        await db.raporlar.create_index("rapor_no", unique=True)
        await db.raporlar.create_index("kategori")
        await db.raporlar.create_index("created_at")
        await db.raporlar.create_index("gecerlilik_tarihi")
        await db.raporlar.create_index("uygunluk")
        await db.raporlar.create_index([("created_at", -1)])  # Descending for latest first
        
        # Kategoriler collection indexes
        await db.kategoriler.create_index("isim", unique=True)
        
        logger.info("Database indexes created successfully")
    except Exception as e:
        logger.warning(f"Index creation error (may already exist): {e}")
    
    # Create default admin if not exists
    admin_exists = await db.users.find_one({"email": "admin@ekipman.com"})
    if not admin_exists:
        admin = User(
            username="admin",
            email="admin@ekipman.com",
            password=get_password_hash("admin123"),
            role="admin",
            email_verified=True
        )
        doc = admin.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.users.insert_one(doc)
        logger.info("Default admin created")
    
    # Create default categories with subcategories
    for cat_name, alt_kats in KATEGORI_ALT_KATEGORI.items():
        exists = await db.kategoriler.find_one({"isim": cat_name})
        if not exists:
            kategori = Kategori(
                isim=cat_name,
                alt_kategoriler=alt_kats,
                aciklama=f"{cat_name} ekipmanları"
            )
            doc = kategori.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            await db.kategoriler.insert_one(doc)
    
    # Create default project
    default_proje_exists = await db.projeler.find_one({"proje_adi": "Çukurova Deprem Konutları Projesi"})
    if not default_proje_exists:
        default_proje = Proje(
            proje_adi="Çukurova Deprem Konutları Projesi",
            aciklama="Varsayılan proje"
        )
        doc = default_proje.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.projeler.insert_one(doc)
        logger.info("Default project created")
        
        # Assign all existing reports to this project
        default_proje_id = doc["id"]
        await db.raporlar.update_many(
            {"proje_id": {"$exists": False}},
            {"$set": {
                "proje_id": default_proje_id,
                "proje_adi": "Çukurova Deprem Konutları Projesi",
                "sehir": "Adana",
                "sehir_kodu": "ADA"
            }}
        )
        logger.info("Existing reports assigned to default project")

# Auth Routes
@api_router.post("/auth/register", response_model=UserResponse)
async def register(user_create: UserCreate):
    # Validate passwords match
    if user_create.password != user_create.password_confirm:
        raise HTTPException(status_code=400, detail="Şifreler eşleşmiyor")
    
    # Check password length
    if len(user_create.password) < 6:
        raise HTTPException(status_code=400, detail="Şifre en az 6 karakter olmalıdır")
    
    # Validate firma exists in database
    firma_exists = await db.raporlar.find_one({"firma": user_create.firma_adi}, {"_id": 0, "firma": 1})
    if not firma_exists:
        raise HTTPException(
            status_code=404, 
            detail="FIRMA_NOT_FOUND"  # Special code for frontend to handle
        )
    
    # Check if email exists
    existing_email = await db.users.find_one({"email": user_create.email})
    if existing_email:
        raise HTTPException(status_code=400, detail="Bu email zaten kayıtlı")
    
    # Check if username exists
    existing_username = await db.users.find_one({"username": user_create.username})
    if existing_username:
        raise HTTPException(status_code=400, detail="Bu kullanıcı adı zaten alınmış")
    
    # Generate verification code
    import random
    verification_code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
    
    user = User(
        username=user_create.username,
        email=user_create.email,
        password=get_password_hash(user_create.password),
        role="viewer",  # Always viewer for self-registration
        firma_adi=user_create.firma_adi,
        email_verified=False,
        verification_code=verification_code
    )
    
    doc = user.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    
    # In production, send email here
    # For now, just log the code
    logger.info(f"Verification code for {user.email}: {verification_code}")
    
    return UserResponse(
        id=user.id,
        username=user.username,
        email=user.email,
        role=user.role,
        firma_adi=user.firma_adi,
        email_verified=user.email_verified,
        created_at=user.created_at
    )

@api_router.post("/auth/verify-email")
async def verify_email(verify_data: VerifyEmail):
    user = await db.users.find_one({"email": verify_data.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    
    if user.get("email_verified"):
        raise HTTPException(status_code=400, detail="Email zaten doğrulanmış")
    
    if user.get("verification_code") != verify_data.code:
        raise HTTPException(status_code=400, detail="Doğrulama kodu hatalı")
    
    await db.users.update_one(
        {"email": verify_data.email},
        {"$set": {"email_verified": True, "verification_code": None}}
    )
    
    return {"message": "Email başarıyla doğrulandı"}

@api_router.post("/auth/resend-code")
async def resend_verification_code(email: EmailStr):
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    
    if user.get("email_verified"):
        raise HTTPException(status_code=400, detail="Email zaten doğrulanmış")
    
    # Generate new code
    import random
    verification_code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
    
    await db.users.update_one(
        {"email": email},
        {"$set": {"verification_code": verification_code}}
    )
    
    logger.info(f"New verification code for {email}: {verification_code}")
    
    return {"message": "Doğrulama kodu yeniden gönderildi"}

@api_router.post("/auth/login", response_model=Token)
async def login(user_login: UserLogin):
    user = await db.users.find_one({"email": user_login.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Email veya şifre hatalı")
    
    if not verify_password(user_login.password, user["password"]):
        raise HTTPException(status_code=401, detail="Email veya şifre hatalı")
    
    access_token = create_access_token(data={"sub": user["id"]})
    
    return Token(
        access_token=access_token,
        token_type="bearer",
        user=UserResponse(
            id=user["id"],
            username=user["username"],
            email=user["email"],
            role=user["role"],
            firma_adi=user.get("firma_adi"),
            email_verified=user.get("email_verified", False),
            created_at=datetime.fromisoformat(user["created_at"]) if isinstance(user["created_at"], str) else user["created_at"]
        )
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(
        id=current_user["id"],
        username=current_user["username"],
        email=current_user["email"],
        role=current_user["role"],
        firma_adi=current_user.get("firma_adi"),
        email_verified=current_user.get("email_verified", False),
        created_at=datetime.fromisoformat(current_user["created_at"]) if isinstance(current_user["created_at"], str) else current_user["created_at"]
    )

# Kategori Routes
@api_router.get("/kategoriler", response_model=List[Kategori])
async def get_kategoriler(current_user: dict = Depends(get_current_user)):
    kategoriler = await db.kategoriler.find({}, {"_id": 0}).to_list(1000)
    for kat in kategoriler:
        if isinstance(kat['created_at'], str):
            kat['created_at'] = datetime.fromisoformat(kat['created_at'])
    return kategoriler

@api_router.post("/kategoriler", response_model=Kategori)
async def create_kategori(kategori_create: KategoriCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    kategori = Kategori(**kategori_create.model_dump())
    doc = kategori.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.kategoriler.insert_one(doc)
    return kategori

@api_router.put("/kategoriler/{kategori_id}")
async def update_kategori(kategori_id: str, kategori_update: KategoriCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    existing = await db.kategoriler.find_one({"id": kategori_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Kategori bulunamadı")
    
    update_data = kategori_update.model_dump()
    await db.kategoriler.update_one({"id": kategori_id}, {"$set": update_data})
    
    updated_kategori = await db.kategoriler.find_one({"id": kategori_id}, {"_id": 0})
    return updated_kategori

@api_router.delete("/kategoriler/{kategori_id}")
async def delete_kategori(kategori_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    result = await db.kategoriler.delete_one({"id": kategori_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kategori bulunamadı")
    return {"message": "Kategori silindi"}

@api_router.post("/kategoriler/bulk-delete")
async def bulk_delete_kategoriler(kategori_ids: List[str], current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    result = await db.kategoriler.delete_many({"id": {"$in": kategori_ids}})
    return {"message": f"{result.deleted_count} kategori silindi", "deleted_count": result.deleted_count}

# Rapor Routes
@api_router.get("/raporlar", response_model=List[Rapor])
async def get_raporlar(
    arama: Optional[str] = None,
    kategori: Optional[str] = None,
    periyot: Optional[str] = None,
    uygunluk: Optional[str] = None,
    firma: Optional[str] = None,
    limit: int = 500,
    skip: int = 0,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    
    # Firma-based access control for viewers
    user_firma = current_user.get("firma_adi")
    if user_firma and current_user.get("role") == "viewer":
        # Viewers can only see their own company's reports
        query["firma"] = user_firma
    elif firma:
        # Admin/Inspector can filter by firma
        query["firma"] = firma
    
    if arama:
        # If user has firma restriction, add it to $and
        if user_firma and current_user.get("role") == "viewer":
            query["$and"] = [
                {"firma": user_firma},
                {"$or": [
                    {"rapor_no": {"$regex": arama, "$options": "i"}},
                    {"ekipman_adi": {"$regex": arama, "$options": "i"}},
                    {"firma": {"$regex": arama, "$options": "i"}}
                ]}
            ]
        else:
            query["$or"] = [
                {"rapor_no": {"$regex": arama, "$options": "i"}},
                {"ekipman_adi": {"$regex": arama, "$options": "i"}},
                {"firma": {"$regex": arama, "$options": "i"}}
            ]
    
    if kategori:
        query["kategori"] = kategori
    
    if periyot:
        query["periyot"] = periyot
    
    if uygunluk:
        query["uygunluk"] = uygunluk
    
    # Limit to 500 records by default for better performance
    raporlar = await db.raporlar.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    for rapor in raporlar:
        if isinstance(rapor['created_at'], str):
            rapor['created_at'] = datetime.fromisoformat(rapor['created_at'])
        if isinstance(rapor['updated_at'], str):
            rapor['updated_at'] = datetime.fromisoformat(rapor['updated_at'])
        # Fallback for old reports without username
        if 'created_by_username' not in rapor or not rapor['created_by_username']:
            rapor['created_by_username'] = 'Bilinmiyor'
    
    return raporlar

@api_router.get("/raporlar/{rapor_id}", response_model=Rapor)
async def get_rapor(rapor_id: str, current_user: dict = Depends(get_current_user)):
    rapor = await db.raporlar.find_one({"id": rapor_id}, {"_id": 0})
    if not rapor:
        raise HTTPException(status_code=404, detail="Rapor bulunamadı")
    
    if isinstance(rapor['created_at'], str):
        rapor['created_at'] = datetime.fromisoformat(rapor['created_at'])
    if isinstance(rapor['updated_at'], str):
        rapor['updated_at'] = datetime.fromisoformat(rapor['updated_at'])
    # Fallback for old reports
    if 'created_by_username' not in rapor or not rapor['created_by_username']:
        rapor['created_by_username'] = 'Bilinmiyor'
    
    return rapor

@api_router.post("/raporlar", response_model=Rapor)
async def create_rapor(rapor_create: RaporCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="Rapor oluşturma yetkiniz yok")
    
    # Get proje
    proje = await db.projeler.find_one({"id": rapor_create.proje_id}, {"_id": 0})
    if not proje:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    # Get sehir kodu
    sehir_obj = next((s for s in SEHIRLER if s["isim"] == rapor_create.sehir), None)
    if not sehir_obj:
        raise HTTPException(status_code=400, detail="Geçersiz şehir")
    
    rapor_no = await generate_rapor_no(rapor_create.sehir)
    
    rapor = Rapor(
        rapor_no=rapor_no,
        proje_id=proje["id"],
        proje_adi=proje["proje_adi"],
        sehir=rapor_create.sehir,
        sehir_kodu=sehir_obj["kod"],
        created_by=current_user["id"],
        created_by_username=current_user["username"],
        ekipman_adi=rapor_create.ekipman_adi,
        kategori=rapor_create.kategori,
        alt_kategori=rapor_create.alt_kategori,
        firma=rapor_create.firma,
        lokasyon=rapor_create.lokasyon,
        marka_model=rapor_create.marka_model,
        seri_no=rapor_create.seri_no,
        periyot=rapor_create.periyot,
        gecerlilik_tarihi=rapor_create.gecerlilik_tarihi,
        aciklama=rapor_create.aciklama,
        uygunluk=rapor_create.uygunluk
    )
    
    doc = rapor.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.raporlar.insert_one(doc)
    
    return rapor

@api_router.put("/raporlar/{rapor_id}", response_model=Rapor)
async def update_rapor(
    rapor_id: str,
    rapor_update: RaporUpdate,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="Rapor düzenleme yetkiniz yok")
    
    rapor = await db.raporlar.find_one({"id": rapor_id}, {"_id": 0})
    if not rapor:
        raise HTTPException(status_code=404, detail="Rapor bulunamadı")
    
    update_data = {k: v for k, v in rapor_update.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.raporlar.update_one({"id": rapor_id}, {"$set": update_data})
    
    updated_rapor = await db.raporlar.find_one({"id": rapor_id}, {"_id": 0})
    if isinstance(updated_rapor['created_at'], str):
        updated_rapor['created_at'] = datetime.fromisoformat(updated_rapor['created_at'])
    if isinstance(updated_rapor['updated_at'], str):
        updated_rapor['updated_at'] = datetime.fromisoformat(updated_rapor['updated_at'])
    
    return updated_rapor

@api_router.delete("/raporlar/{rapor_id}")
async def delete_rapor(rapor_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="Rapor silme yetkiniz yok")
    
    # Delete associated files
    dosyalar = await db.medya_dosyalari.find({"rapor_id": rapor_id}, {"_id": 0}).to_list(100)
    for dosya in dosyalar:
        dosya_path = Path(dosya["dosya_yolu"])
        if dosya_path.exists():
            dosya_path.unlink()
    
    await db.medya_dosyalari.delete_many({"rapor_id": rapor_id})
    
    result = await db.raporlar.delete_one({"id": rapor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Rapor bulunamadı")
    
    return {"message": "Rapor silindi"}

@api_router.patch("/raporlar/{rapor_id}/durum")
async def toggle_rapor_durum(rapor_id: str, current_user: dict = Depends(get_current_user)):
    """Toggle report status between Aktif and Pasif"""
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="Rapor durumunu değiştirme yetkiniz yok")
    
    rapor = await db.raporlar.find_one({"id": rapor_id}, {"_id": 0})
    if not rapor:
        raise HTTPException(status_code=404, detail="Rapor bulunamadı")
    
    # Toggle durum
    yeni_durum = "Pasif" if rapor.get("durum", "Aktif") == "Aktif" else "Aktif"
    
    await db.raporlar.update_one(
        {"id": rapor_id},
        {"$set": {"durum": yeni_durum, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": f"Rapor durumu {yeni_durum} olarak güncellendi", "durum": yeni_durum}

@api_router.post("/raporlar/bulk-delete")
async def bulk_delete_raporlar(rapor_ids: List[str], current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="Rapor silme yetkiniz yok")
    
    # Delete associated files for all reports
    for rapor_id in rapor_ids:
        dosyalar = await db.medya_dosyalari.find({"rapor_id": rapor_id}, {"_id": 0}).to_list(100)
        for dosya in dosyalar:
            dosya_path = Path(dosya["dosya_yolu"])
            if dosya_path.exists():
                dosya_path.unlink()
        await db.medya_dosyalari.delete_many({"rapor_id": rapor_id})
    
    result = await db.raporlar.delete_many({"id": {"$in": rapor_ids}})
    return {"message": f"{result.deleted_count} rapor silindi", "deleted_count": result.deleted_count}

# ZIP Export Route - Seçili raporları ZIP olarak indir
class ZipExportRequest(BaseModel):
    rapor_ids: List[str]

@api_router.post("/raporlar/zip-export")
async def zip_export_raporlar(
    request: ZipExportRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Seçilen raporları klasör yapısıyla ZIP dosyası olarak indirir.
    Her rapor için ayrı klasör oluşturulur ve içine rapor dosyaları ile bilgi.txt eklenir.
    """
    rapor_ids = request.rapor_ids
    
    if not rapor_ids:
        raise HTTPException(status_code=400, detail="En az bir rapor seçilmelidir")
    
    if len(rapor_ids) > 100:
        raise HTTPException(status_code=400, detail="En fazla 100 rapor seçilebilir")
    
    # Seçilen raporları getir
    raporlar = await db.raporlar.find({"id": {"$in": rapor_ids}}, {"_id": 0}).to_list(100)
    
    if not raporlar:
        raise HTTPException(status_code=404, detail="Seçilen raporlar bulunamadı")
    
    # Geçici klasör oluştur
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Her rapor için klasör oluştur
        for rapor in raporlar:
            rapor_no = rapor.get("rapor_no", f"RAPOR_{rapor.get('id', 'unknown')[:8]}")
            # Klasör adını güvenli hale getir (özel karakterleri kaldır)
            safe_rapor_no = "".join(c if c.isalnum() or c in "-_" else "_" for c in rapor_no)
            rapor_folder = os.path.join(temp_dir, f"RAPOR_{safe_rapor_no}")
            os.makedirs(rapor_folder, exist_ok=True)
            
            # bilgi.txt dosyası oluştur
            bilgi_content = f"""╔══════════════════════════════════════════════════════════════╗
║                    RAPOR BİLGİLERİ                           ║
╚══════════════════════════════════════════════════════════════╝

📋 Rapor No        : {rapor.get('rapor_no', 'Belirtilmemiş')}
📅 Oluşturma Tarihi: {rapor.get('created_at', 'Belirtilmemiş')[:10] if rapor.get('created_at') else 'Belirtilmemiş'}
🏢 Firma           : {rapor.get('firma', 'Belirtilmemiş')}
🔧 Ekipman Adı     : {rapor.get('ekipman_adi', 'Belirtilmemiş')}
📂 Kategori        : {rapor.get('kategori', 'Belirtilmemiş')}
📁 Alt Kategori    : {rapor.get('alt_kategori', 'Belirtilmemiş')}
📍 Lokasyon        : {rapor.get('lokasyon', 'Belirtilmemiş')}
🏭 Marka/Model     : {rapor.get('marka_model', 'Belirtilmemiş')}
🔢 Seri No         : {rapor.get('seri_no', 'Belirtilmemiş')}
⏱️ Periyot         : {rapor.get('periyot', 'Belirtilmemiş')}
📅 Geçerlilik      : {rapor.get('gecerlilik_tarihi', 'Belirtilmemiş')}
✅ Uygunluk        : {rapor.get('uygunluk', 'Belirtilmemiş')}
🏙️ Şehir           : {rapor.get('sehir', 'Belirtilmemiş')}
📝 Proje           : {rapor.get('proje_adi', 'Belirtilmemiş')}
👤 Oluşturan       : {rapor.get('created_by_username', 'Belirtilmemiş')}
📊 Durum           : {rapor.get('durum', 'Aktif')}

═══════════════════════════════════════════════════════════════
📝 AÇIKLAMA:
───────────────────────────────────────────────────────────────
{rapor.get('aciklama', 'Açıklama bulunmamaktadır.')}
═══════════════════════════════════════════════════════════════

Bu dosya EKOS - Ekipman Kontrol Otomasyon Sistemi tarafından 
otomatik olarak oluşturulmuştur.
Tarih: {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M:%S')} UTC
"""
            
            bilgi_path = os.path.join(rapor_folder, "bilgi.txt")
            with open(bilgi_path, "w", encoding="utf-8") as f:
                f.write(bilgi_content)
            
            # Rapora ait dosyaları kopyala
            rapor_id = rapor.get("id")
            dosyalar = await db.medya_dosyalari.find({"rapor_id": rapor_id}, {"_id": 0}).to_list(100)
            
            for idx, dosya in enumerate(dosyalar):
                dosya_path = Path(dosya.get("dosya_yolu", ""))
                if dosya_path.exists():
                    # Orijinal dosya adını kullan
                    original_name = dosya.get("dosya_adi", f"dosya_{idx}")
                    # Dosya adını güvenli hale getir
                    safe_name = "".join(c if c.isalnum() or c in ".-_" else "_" for c in original_name)
                    dest_path = os.path.join(rapor_folder, safe_name)
                    
                    # Aynı isimde dosya varsa numara ekle
                    counter = 1
                    base_name, ext = os.path.splitext(safe_name)
                    while os.path.exists(dest_path):
                        dest_path = os.path.join(rapor_folder, f"{base_name}_{counter}{ext}")
                        counter += 1
                    
                    shutil.copy2(str(dosya_path), dest_path)
        
        # ZIP dosyası oluştur
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zip_file.write(file_path, arcname)
        
        zip_buffer.seek(0)
        
        # Dosya adı oluştur
        now = datetime.now(timezone.utc)
        username = current_user.get("username", "user")
        zip_filename = f"Raporlar_{username}_{now.strftime('%Y%m%d_%H%M')}.zip"
        
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{zip_filename}"',
                "Content-Type": "application/zip"
            }
        )
        
    finally:
        # Geçici klasörü temizle
        shutil.rmtree(temp_dir, ignore_errors=True)

# File Upload Routes
@api_router.post("/upload/{rapor_id}")
async def upload_file(
    rapor_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="Dosya yükleme yetkiniz yok")
    
    # Check file size (4GB)
    content = await file.read()
    if len(content) > 4 * 1024 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Dosya boyutu 4GB'dan büyük olamaz")
    
    # Check file type
    allowed_types = ["image/jpeg", "image/jpg", "image/png", "application/pdf"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Sadece JPG, PNG ve PDF formatları desteklenir")
    
    # Check if report exists
    rapor = await db.raporlar.find_one({"id": rapor_id})
    if not rapor:
        raise HTTPException(status_code=404, detail="Rapor bulunamadı")
    
    # Save file
    file_ext = file.filename.split(".")[-1]
    file_id = str(uuid.uuid4())
    filename = f"{file_id}.{file_ext}"
    file_path = UPLOAD_DIR / filename
    
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Save to database
    medya = MedyaDosyasi(
        rapor_id=rapor_id,
        dosya_adi=file.filename,
        dosya_yolu=str(file_path),
        dosya_tipi=file.content_type,
        dosya_boyutu=len(content)
    )
    
    doc = medya.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.medya_dosyalari.insert_one(doc)
    
    return {"message": "Dosya yüklendi", "file_id": medya.id}

@api_router.get("/dosyalar/{rapor_id}")
async def get_dosyalar(rapor_id: str, current_user: dict = Depends(get_current_user)):
    dosyalar = await db.medya_dosyalari.find({"rapor_id": rapor_id}, {"_id": 0}).to_list(100)
    for dosya in dosyalar:
        if isinstance(dosya['created_at'], str):
            dosya['created_at'] = datetime.fromisoformat(dosya['created_at'])
    return dosyalar

@api_router.get("/dosyalar/{dosya_id}/indir")
async def download_dosya(dosya_id: str, current_user: dict = Depends(get_current_user)):
    dosya = await db.medya_dosyalari.find_one({"id": dosya_id}, {"_id": 0})
    if not dosya:
        raise HTTPException(status_code=404, detail="Dosya bulunamadı")
    
    dosya_path = Path(dosya["dosya_yolu"])
    if not dosya_path.exists():
        raise HTTPException(status_code=404, detail="Dosya sistemde bulunamadı")
    
    from fastapi.responses import FileResponse
    
    # Determine correct media type
    file_ext = dosya["dosya_adi"].lower().split('.')[-1]
    media_type_map = {
        'pdf': 'application/pdf',
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'png': 'image/png',
        'gif': 'image/gif',
        'webp': 'image/webp',
        'bmp': 'image/bmp',
        'svg': 'image/svg+xml',
    }
    
    media_type = media_type_map.get(file_ext, dosya.get("dosya_tipi", "application/octet-stream"))
    
    return FileResponse(
        path=str(dosya_path),
        filename=dosya["dosya_adi"],
        media_type=media_type,
        headers={
            "Content-Disposition": f'inline; filename="{dosya["dosya_adi"]}"',
            "Cache-Control": "public, max-age=3600"
        }
    )

@api_router.delete("/dosyalar/{dosya_id}")
async def delete_dosya(dosya_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="Dosya silme yetkiniz yok")
    
    dosya = await db.medya_dosyalari.find_one({"id": dosya_id}, {"_id": 0})
    if not dosya:
        raise HTTPException(status_code=404, detail="Dosya bulunamadı")
    
    dosya_path = Path(dosya["dosya_yolu"])
    if dosya_path.exists():
        dosya_path.unlink()
    
    await db.medya_dosyalari.delete_one({"id": dosya_id})
    return {"message": "Dosya silindi"}

# Excel Routes
@api_router.get("/excel/export")
async def export_excel(current_user: dict = Depends(get_current_user)):
    raporlar = await db.raporlar.find({}, {"_id": 0}).to_list(10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Raporlar"
    
    # Headers
    headers = [
        "Rapor No", "Ekipman Adı", "Kategori", "Firma", "Lokasyon",
        "Marka/Model", "Seri No", "Alt Kategori", "Periyot",
        "Geçerlilik Tarihi", "Uygunluk", "Açıklama", "Oluşturma Tarihi"
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data
    for row_idx, rapor in enumerate(raporlar, 2):
        ws.cell(row=row_idx, column=1, value=rapor.get("rapor_no", ""))
        ws.cell(row=row_idx, column=2, value=rapor.get("ekipman_adi", ""))
        ws.cell(row=row_idx, column=3, value=rapor.get("kategori", ""))
        ws.cell(row=row_idx, column=4, value=rapor.get("firma", ""))
        ws.cell(row=row_idx, column=5, value=rapor.get("lokasyon", ""))
        ws.cell(row=row_idx, column=6, value=rapor.get("marka_model", ""))
        ws.cell(row=row_idx, column=7, value=rapor.get("seri_no", ""))
        ws.cell(row=row_idx, column=8, value=rapor.get("alt_kategori", ""))
        ws.cell(row=row_idx, column=9, value=rapor.get("periyot", ""))
        ws.cell(row=row_idx, column=10, value=rapor.get("gecerlilik_tarihi", ""))
        ws.cell(row=row_idx, column=11, value=rapor.get("uygunluk", ""))
        ws.cell(row=row_idx, column=12, value=rapor.get("aciklama", ""))
        created_at = rapor.get("created_at", "")
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        ws.cell(row=row_idx, column=13, value=created_at.strftime("%Y-%m-%d %H:%M") if created_at else "")
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=raporlar.xlsx"}
    )

@api_router.get("/excel/template")
async def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapor Şablonu"
    
    # Updated headers with Şehir as first column
    headers = [
        "Şehir", "Ekipman Adı", "Kategori", "Firma", "Lokasyon", "Marka/Model",
        "Seri No", "Alt Kategori", "Periyot", "Geçerlilik Tarihi",
        "Uygunluk", "Açıklama"
    ]
    
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Example rows with different cities
    examples = [
        ["İstanbul", "Asansör A1", "Asansör", "ABC Firma", "İstanbul Ofis", "Otis 2000",
         "SN12345", "Yolcu Asansörü", "6 Aylık", "2025-12-31", "Uygun", "Örnek açıklama"],
        ["Ankara", "Hidrofor H2", "Basınçlı Kaplar", "XYZ Ltd", "Ankara Fabrika", "Grundfos",
         "SN67890", "Hidrofor", "3 Aylık", "2025-09-30", "Uygun Değil", ""],
        ["İzmir", "Forklift F3", "Forklift", "Test Firma", "İzmir Depo", "Toyota", "", "", "", "", "", ""]
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col, value in enumerate(example, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    # Add city list in a separate sheet
    ws_cities = wb.create_sheet("Şehir Listesi")
    ws_cities.cell(row=1, column=1, value="Geçerli Şehirler (Büyük/küçük harf önemli değil)")
    ws_cities.cell(row=1, column=1).font = Font(bold=True, size=12)
    
    for idx, sehir in enumerate(SEHIRLER, 2):
        ws_cities.cell(row=idx, column=1, value=sehir["isim"])
    
    ws_cities.column_dimensions['A'].width = 25
    
    # Set column widths for main sheet
    for col in ws.columns:
        column = col[0].column_letter
        ws.column_dimensions[column].width = 20
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rapor_sablonu.xlsx"}
    )

@api_router.post("/excel/import")
async def import_excel(
    file: UploadFile = File(...),
    proje_id: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="Excel içe aktarma yetkiniz yok")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sadece Excel dosyaları yüklenebilir")
    
    # Validate and get project details
    proje = await db.projeler.find_one({"id": proje_id}, {"_id": 0})
    if not proje:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    proje_adi = proje.get("proje_adi", "")
    
    content = await file.read()
    excel_file = io.BytesIO(content)
    
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        
        imported_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):  # Skip empty rows
                continue
            
            try:
                # Excel format: Şehir | Ekipman Adı | Kategori | Firma | Lokasyon | Marka/Model | Seri No | Alt Kategori | Periyot | Geçerlilik Tarihi | Uygunluk | Açıklama
                # Safely get values with default None if index doesn't exist
                def get_cell(index):
                    return row[index] if index < len(row) and row[index] is not None else None
                
                sehir = str(get_cell(0)) if get_cell(0) else None
                ekipman_adi = str(get_cell(1)) if get_cell(1) else ""
                kategori = str(get_cell(2)) if get_cell(2) else ""
                firma = str(get_cell(3)) if get_cell(3) else ""
                lokasyon = str(get_cell(4)) if get_cell(4) else None
                marka_model = str(get_cell(5)) if get_cell(5) else None
                seri_no = str(get_cell(6)) if get_cell(6) else None
                alt_kategori = str(get_cell(7)) if get_cell(7) else None
                periyot = str(get_cell(8)) if get_cell(8) else None
                gecerlilik_tarihi = str(get_cell(9)) if get_cell(9) else None
                uygunluk = str(get_cell(10)) if get_cell(10) else None
                aciklama = str(get_cell(11)) if get_cell(11) else None
                
                # Validate required fields
                if not ekipman_adi or not kategori or not firma:
                    errors.append(f"Satır {row_idx}: Zorunlu alanlar eksik (Ekipman Adı, Kategori, Firma)")
                    continue
                
                if not sehir:
                    errors.append(f"Satır {row_idx}: Şehir alanı zorunludur")
                    continue
                
                # Get sehir kodu with case-insensitive and Turkish character normalization
                # Turkish locale-aware lowercase conversion
                def turkish_lower(text):
                    text = text.strip()
                    # First handle Turkish specific characters
                    text = text.replace('İ', 'i').replace('I', 'ı')
                    # Then apply standard lower
                    return text.lower()
                
                def normalize_turkish(text):
                    # Normalize Turkish characters to ASCII for comparison
                    turkish_map = str.maketrans('ıİiIğĞüÜşŞöÖçÇ', 'iiiigguussoocc')
                    return turkish_lower(text).translate(turkish_map)
                
                sehir_normalized = normalize_turkish(sehir)
                
                sehir_obj = None
                for s in SEHIRLER:
                    if normalize_turkish(s["isim"]) == sehir_normalized:
                        sehir_obj = s
                        sehir = s["isim"]  # Use the correct city name
                        break
                
                if not sehir_obj:
                    # Provide helpful error message with available cities
                    errors.append(f"Satır {row_idx}: Geçersiz şehir - '{sehir}'. Lütfen geçerli bir Türkiye şehri adı girin (örn: İstanbul, Ankara, İzmir)")
                    continue
                
                # Generate rapor no with city
                rapor_no = await generate_rapor_no(sehir)
                
                rapor_data = {
                    "rapor_no": rapor_no,
                    "proje_id": proje_id,
                    "proje_adi": proje_adi,
                    "sehir": sehir,
                    "sehir_kodu": sehir_obj["kod"],
                    "ekipman_adi": ekipman_adi,
                    "kategori": kategori,
                    "alt_kategori": alt_kategori,
                    "firma": firma,
                    "lokasyon": lokasyon,
                    "marka_model": marka_model,
                    "seri_no": seri_no,
                    "periyot": periyot,
                    "gecerlilik_tarihi": gecerlilik_tarihi,
                    "uygunluk": uygunluk,
                    "aciklama": aciklama,
                    "durum": "Aktif",
                    "created_by": current_user["id"],
                    "created_by_username": current_user.get("username", current_user.get("email", ""))
                }
                
                rapor = Rapor(**rapor_data)
                doc = rapor.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                doc['updated_at'] = doc['updated_at'].isoformat()
                await db.raporlar.insert_one(doc)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Satır {row_idx}: {str(e)}")
        
        return {
            "message": f"{imported_count} rapor başarıyla içe aktarıldı",
            "imported_count": imported_count,
            "errors": errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel işleme hatası: {str(e)}")

# Dashboard Stats
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    # Only admin and inspector can access dashboard
    if current_user.get("role") not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="Dashboard'a erişim yetkiniz yok")
    
    # Firma-based filter for viewers (keeping for future use)
    base_query = {}
    user_firma = current_user.get("firma_adi")
    if user_firma and current_user.get("role") == "viewer":
        base_query["firma"] = user_firma
    
    # Total reports
    total_raporlar = await db.raporlar.count_documents(base_query)
    
    # This month reports
    now = datetime.now(timezone.utc)
    start_of_month = datetime(now.year, now.month, 1, tzinfo=timezone.utc)
    monthly_query = {**base_query, "created_at": {"$gte": start_of_month.isoformat()}}
    monthly_raporlar = await db.raporlar.count_documents(monthly_query)
    
    # Uygunluk stats
    uygun_count = await db.raporlar.count_documents({**base_query, "uygunluk": "Uygun"})
    uygun_degil_count = await db.raporlar.count_documents({**base_query, "uygunluk": "Uygun Değil"})
    
    # Expiring reports (30 days) - optimized query
    now_date = now.date()
    thirty_days = (now + timedelta(days=30)).date()
    seven_days = (now + timedelta(days=7)).date()
    
    # Only fetch gecerlilik_tarihi field for performance
    date_query = {**base_query, "gecerlilik_tarihi": {"$ne": None, "$ne": ""}}
    raporlar_with_dates = await db.raporlar.find(
        date_query, 
        {"gecerlilik_tarihi": 1, "_id": 0}
    ).limit(5000).to_list(5000)
    
    expiring_30_days = 0
    expiring_7_days = 0
    
    for rapor in raporlar_with_dates:
        gecerlilik_str = rapor.get("gecerlilik_tarihi")
        if gecerlilik_str and str(gecerlilik_str).strip():
            try:
                from datetime import datetime as dt
                # Try multiple date formats
                gecerlilik = None
                for fmt in ["%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"]:
                    try:
                        gecerlilik = dt.strptime(str(gecerlilik_str), fmt).date()
                        break
                    except:
                        continue
                
                if gecerlilik:
                    if now_date <= gecerlilik <= thirty_days:
                        expiring_30_days += 1
                    if now_date <= gecerlilik <= seven_days:
                        expiring_7_days += 1
            except Exception as e:
                continue
    
    # Category distribution (top 6) - get unique categories from reports
    pipeline = []
    if base_query:
        pipeline.append({"$match": base_query})
    pipeline.extend([
        {"$group": {"_id": "$kategori", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 6},
        {"$project": {"kategori": "$_id", "count": 1, "_id": 0}}
    ])
    
    kategori_dagilim = await db.raporlar.aggregate(pipeline).to_list(6)
    
    # İskele Bileşenleri İstatistikleri
    iskele_query = {}
    if current_user.get("role") == "viewer" and current_user.get("firma_adi"):
        iskele_query["firma_adi"] = current_user.get("firma_adi")
    
    # Toplam bileşen adedi (bileşen_adedi alanlarının toplamı)
    total_pipeline = []
    if iskele_query:
        total_pipeline.append({"$match": iskele_query})
    total_pipeline.append({"$group": {"_id": None, "total": {"$sum": "$bileşen_adedi"}}})
    
    total_result = await db.iskele_bilesenleri.aggregate(total_pipeline).to_list(1)
    total_iskele = total_result[0]["total"] if total_result else 0
    
    # Uygun olanların bileşen adedi toplamı
    uygun_pipeline = []
    if iskele_query:
        uygun_pipeline.append({"$match": iskele_query})
    uygun_pipeline.extend([
        {"$match": {"uygunluk": {"$regex": "^uygun$", "$options": "i"}}},
        {"$group": {"_id": None, "total": {"$sum": "$bileşen_adedi"}}}
    ])
    
    uygun_result = await db.iskele_bilesenleri.aggregate(uygun_pipeline).to_list(1)
    iskele_uygun = uygun_result[0]["total"] if uygun_result else 0
    
    # Uygun olmayanların bileşen adedi toplamı - case-insensitive regex
    uygun_degil_pipeline = []
    if iskele_query:
        uygun_degil_pipeline.append({"$match": iskele_query})
    uygun_degil_pipeline.extend([
        {"$match": {"uygunluk": {"$regex": "uygun.*(de[ğg]il|olmayan)", "$options": "i"}}},
        {"$group": {"_id": None, "total": {"$sum": "$bileşen_adedi"}}}
    ])
    
    uygun_degil_result = await db.iskele_bilesenleri.aggregate(uygun_degil_pipeline).to_list(1)
    iskele_uygun_degil = uygun_degil_result[0]["total"] if uygun_degil_result else 0
    
    # Bileşen adlarına göre dağılım (bileşen adetlerinin toplamı)
    iskele_pipeline = []
    if iskele_query:
        iskele_pipeline.append({"$match": iskele_query})
    iskele_pipeline.extend([
        {"$group": {"_id": "$bileşen_adi", "count": {"$sum": "$bileşen_adedi"}}},
        {"$sort": {"count": -1}},
        {"$limit": 6},
        {"$project": {"bileşen_adi": "$_id", "count": 1, "_id": 0}}
    ])
    
    bilesen_dagilim = await db.iskele_bilesenleri.aggregate(iskele_pipeline).to_list(6)
    
    return {
        "total_raporlar": total_raporlar,
        "monthly_raporlar": monthly_raporlar,
        "uygun_count": uygun_count,
        "uygun_degil_count": uygun_degil_count,
        "expiring_30_days": expiring_30_days,
        "expiring_7_days": expiring_7_days,
        "kategori_dagilim": kategori_dagilim,
        "iskele_stats": {
            "total": total_iskele,
            "uygun": iskele_uygun,
            "uygun_degil": iskele_uygun_degil,
            "uygunluk_orani": round((iskele_uygun / total_iskele * 100), 1) if total_iskele > 0 else 0,
            "bilesen_dagilim": bilesen_dagilim
        }
    }

# Users Management (Admin only)
@api_router.get("/users", response_model=List[UserResponse])
async def get_users(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    
    result = []
    for user in users:
        result.append(UserResponse(
            id=user["id"],
            username=user.get("username", user["email"].split("@")[0]),
            email=user["email"],
            role=user["role"],
            email_verified=user.get("email_verified", False),
            created_at=datetime.fromisoformat(user["created_at"]) if isinstance(user["created_at"], str) else user["created_at"]
        ))
    
    return result

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="Kendi hesabınızı silemezsiniz")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    
    return {"message": "Kullanıcı silindi"}

@api_router.post("/users/bulk-delete")
async def bulk_delete_users(user_ids: List[str], current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    # Remove current user from deletion list
    user_ids = [uid for uid in user_ids if uid != current_user["id"]]
    
    if not user_ids:
        raise HTTPException(status_code=400, detail="Silinecek kullanıcı bulunamadı veya kendi hesabınızı silemezsiniz")
    
    result = await db.users.delete_many({"id": {"$in": user_ids}})
    return {"message": f"{result.deleted_count} kullanıcı silindi", "deleted_count": result.deleted_count}

# Şehirler endpoint
@api_router.get("/sehirler")
async def get_sehirler():
    return SEHIRLER

# Kategori-Alt Kategori Map
@api_router.get("/kategori-alt-kategoriler")
async def get_kategori_alt_kategoriler():
    return KATEGORI_ALT_KATEGORI

# Projeler endpoints
@api_router.get("/projeler", response_model=List[Proje])
async def get_projeler(current_user: dict = Depends(get_current_user)):
    projeler = await db.projeler.find({}, {"_id": 0}).to_list(1000)
    for proje in projeler:
        if isinstance(proje['created_at'], str):
            proje['created_at'] = datetime.fromisoformat(proje['created_at'])
    return projeler

@api_router.post("/projeler", response_model=Proje)
async def create_proje(proje_create: ProjeCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    proje = Proje(**proje_create.model_dump())
    doc = proje.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.projeler.insert_one(doc)
    return proje

@api_router.put("/projeler/{proje_id}")
async def update_proje(proje_id: str, proje_update: ProjeCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    existing = await db.projeler.find_one({"id": proje_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    update_data = proje_update.model_dump()
    await db.projeler.update_one({"id": proje_id}, {"$set": update_data})
    
    updated_proje = await db.projeler.find_one({"id": proje_id}, {"_id": 0})
    return updated_proje

@api_router.delete("/projeler/{proje_id}")
async def delete_proje(proje_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    result = await db.projeler.delete_one({"id": proje_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    return {"message": "Proje silindi"}

# ==================== İSKELE BİLEŞEN ADLARI ENDPOINTS ====================

@api_router.get("/iskele-bilesen-adlari")
async def get_iskele_bilesen_adlari(current_user: dict = Depends(get_current_user)):
    bilesen_adlari = await db.iskele_bilesen_adlari.find({}, {"_id": 0}).to_list(1000)
    return bilesen_adlari

@api_router.post("/iskele-bilesen-adlari")
async def create_iskele_bilesen_adi(
    bilesen_adi: str,
    aciklama: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    # Check if already exists
    existing = await db.iskele_bilesen_adlari.find_one({"bilesen_adi": bilesen_adi}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Bu bileşen adı zaten mevcut")
    
    bilesen_id = str(uuid.uuid4())
    bilesen_data = {
        "id": bilesen_id,
        "bilesen_adi": bilesen_adi,
        "aciklama": aciklama,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.iskele_bilesen_adlari.insert_one(bilesen_data)
    created = await db.iskele_bilesen_adlari.find_one({"id": bilesen_id}, {"_id": 0})
    return created

@api_router.put("/iskele-bilesen-adlari/{bilesen_id}")
async def update_iskele_bilesen_adi(
    bilesen_id: str,
    bilesen_adi: str,
    aciklama: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    existing = await db.iskele_bilesen_adlari.find_one({"id": bilesen_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Bileşen adı bulunamadı")
    
    update_data = {
        "bilesen_adi": bilesen_adi,
        "aciklama": aciklama
    }
    
    await db.iskele_bilesen_adlari.update_one({"id": bilesen_id}, {"$set": update_data})
    updated = await db.iskele_bilesen_adlari.find_one({"id": bilesen_id}, {"_id": 0})
    return updated

@api_router.delete("/iskele-bilesen-adlari/{bilesen_id}")
async def delete_iskele_bilesen_adi(
    bilesen_id: str,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    result = await db.iskele_bilesen_adlari.delete_one({"id": bilesen_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bileşen adı bulunamadı")
    
    return {"message": "Bileşen adı silindi"}

# ==================== İSKELE BİLEŞENLERİ ENDPOINTS ====================

@api_router.get("/iskele-bilesenleri")
async def get_iskele_bilesenleri(
    current_user: dict = Depends(get_current_user),
    limit: int = 500
):
    # Role-based filtering
    query = {}
    if current_user.get("role") == "viewer" and current_user.get("firma_adi"):
        query["firma_adi"] = current_user.get("firma_adi")
    
    bilesenleri = await db.iskele_bilesenleri.find(query, {"_id": 0}).to_list(limit)
    return bilesenleri

@api_router.post("/iskele-bilesenleri")
async def create_iskele_bileseni(
    bileşen: IskeleBileseniCreate,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="İskele bileşeni ekleme yetkiniz yok")
    
    # Validate and get project
    proje = await db.projeler.find_one({"id": bileşen.proje_id}, {"_id": 0})
    if not proje:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    bileşen_id = str(uuid.uuid4())
    bileşen_data = {
        **bileşen.model_dump(),
        "id": bileşen_id,
        "proje_adi": proje.get("proje_adi", ""),
        "iskele_periyodu": "6 Aylık",
        "created_by": current_user["id"],
        "created_by_username": current_user.get("username", current_user.get("email", "")),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.iskele_bilesenleri.insert_one(bileşen_data)
    
    # Return without _id
    created = await db.iskele_bilesenleri.find_one({"id": bileşen_id}, {"_id": 0})
    return created

@api_router.get("/iskele-bilesenleri/{bilesen_id}")
async def get_iskele_bileseni(
    bilesen_id: str,
    current_user: dict = Depends(get_current_user)
):
    bileşen = await db.iskele_bilesenleri.find_one({"id": bilesen_id}, {"_id": 0})
    if not bileşen:
        raise HTTPException(status_code=404, detail="İskele bileşeni bulunamadı")
    return bileşen

@api_router.put("/iskele-bilesenleri/{bilesen_id}")
async def update_iskele_bileseni(
    bilesen_id: str,
    bileşen_update: IskeleBileseniCreate,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="İskele bileşeni güncelleme yetkiniz yok")
    
    existing = await db.iskele_bilesenleri.find_one({"id": bilesen_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="İskele bileşeni bulunamadı")
    
    update_data = bileşen_update.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["iskele_periyodu"] = "6 Aylık"  # Always 6 months
    
    await db.iskele_bilesenleri.update_one(
        {"id": bilesen_id},
        {"$set": update_data}
    )
    
    updated = await db.iskele_bilesenleri.find_one({"id": bilesen_id}, {"_id": 0})
    return updated

@api_router.delete("/iskele-bilesenleri/{bilesen_id}")
async def delete_iskele_bileseni(
    bilesen_id: str,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="İskele bileşeni silme yetkiniz yok")
    
    result = await db.iskele_bilesenleri.delete_one({"id": bilesen_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="İskele bileşeni bulunamadı")
    
    return {"message": "İskele bileşeni silindi"}

@api_router.post("/iskele-bilesenleri/bulk-delete")
async def bulk_delete_iskele_bilesenleri_route(
    bilesen_ids: List[str],
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="İskele bileşeni silme yetkiniz yok")
    
    if not bilesen_ids:
        raise HTTPException(status_code=400, detail="Silinecek bileşen ID'leri belirtilmedi")
    
    result = await db.iskele_bilesenleri.delete_many({"id": {"$in": bilesen_ids}})
    return {"message": f"{result.deleted_count} iskele bileşeni silindi", "deleted_count": result.deleted_count}

@api_router.get("/iskele-bilesenleri/excel/export")
async def export_iskele_excel(current_user: dict = Depends(get_current_user)):
    # Role-based filtering
    query = {}
    if current_user.get("role") == "viewer" and current_user.get("firma_adi"):
        query["firma_adi"] = current_user.get("firma_adi")
    
    bilesenleri = await db.iskele_bilesenleri.find(query, {"_id": 0}).to_list(1000)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "İskele Bileşenleri"
    
    # Headers
    headers = [
        "Bileşen Adı", "Malzeme Kodu", "Bileşen Adedi", "Firma Adı",
        "Geçerlilik Tarihi", "Uygunluk", "Açıklama", "Proje Adı"
    ]
    
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for row_idx, bilesen in enumerate(bilesenleri, 2):
        ws.cell(row=row_idx, column=1, value=bilesen.get("bileşen_adi", ""))
        ws.cell(row=row_idx, column=2, value=bilesen.get("malzeme_kodu", ""))
        ws.cell(row=row_idx, column=3, value=bilesen.get("bileşen_adedi", 0))
        ws.cell(row=row_idx, column=4, value=bilesen.get("firma_adi", ""))
        ws.cell(row=row_idx, column=5, value=bilesen.get("gecerlilik_tarihi", ""))
        ws.cell(row=row_idx, column=6, value=bilesen.get("uygunluk", ""))
        ws.cell(row=row_idx, column=7, value=bilesen.get("aciklama", ""))
        ws.cell(row=row_idx, column=8, value=bilesen.get("proje_adi", ""))
    
    # Set column widths
    for col in ws.columns:
        column = col[0].column_letter
        ws.column_dimensions[column].width = 20
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=iskele_bilesenleri_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@api_router.get("/iskele-bilesenleri/excel/template")
async def download_iskele_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "İskele Bileşenleri Şablonu"
    
    # Headers
    headers = [
        "Bileşen Adı", "Malzeme Kodu", "Bileşen Adedi", "Firma Adı",
        "Geçerlilik Tarihi", "Uygunluk", "Açıklama"
    ]
    
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Example rows
    examples = [
        ["Çelik Direk", "CD-001", 10, "ABC İnşaat", "2025-12-31", "Uygun", "Standart çelik direk"],
        ["Bağlantı Elemanı", "BE-002", 50, "XYZ Yapı", "2025-06-30", "Uygun", ""],
        ["Destek Parçası", "DP-003", 25, "Test Firma", "", "Uygun Değil", "Kontrol gerekli"]
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col, value in enumerate(example, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    for col in ws.columns:
        column = col[0].column_letter
        ws.column_dimensions[column].width = 20
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=iskele_bilesenleri_sablonu.xlsx"}
    )

@api_router.post("/iskele-bilesenleri/excel/import")
async def import_iskele_excel(
    file: UploadFile = File(...),
    proje_id: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="İskele bileşeni içe aktarma yetkiniz yok")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sadece Excel dosyaları yüklenebilir")
    
    # Validate and get project
    proje = await db.projeler.find_one({"id": proje_id}, {"_id": 0})
    if not proje:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    proje_adi = proje.get("proje_adi", "")
    
    content = await file.read()
    excel_file = io.BytesIO(content)
    
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        imported_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):  # Skip empty rows
                continue
            
            try:
                # Excel format: Bileşen Adı | Malzeme Kodu | Bileşen Adedi | Firma Adı | Geçerlilik Tarihi | Uygunluk | Açıklama
                def get_cell(index):
                    return row[index] if index < len(row) and row[index] is not None else None
                
                bileşen_adi = str(get_cell(0)) if get_cell(0) else ""
                malzeme_kodu = str(get_cell(1)) if get_cell(1) else ""
                bileşen_adedi_raw = get_cell(2)
                firma_adi = str(get_cell(3)) if get_cell(3) else ""
                gecerlilik_tarihi = str(get_cell(4)) if get_cell(4) else None
                uygunluk = str(get_cell(5)) if get_cell(5) else "Uygun"
                aciklama = str(get_cell(6)) if get_cell(6) else None
                
                # Validate required fields
                if not bileşen_adi or not malzeme_kodu or not firma_adi:
                    errors.append(f"Satır {row_idx}: Zorunlu alanlar eksik (Bileşen Adı, Malzeme Kodu, Firma Adı)")
                    continue
                
                # Parse bileşen_adedi
                try:
                    bileşen_adedi = int(bileşen_adedi_raw) if bileşen_adedi_raw else 1
                    if bileşen_adedi < 1:
                        errors.append(f"Satır {row_idx}: Bileşen adedi en az 1 olmalıdır")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Satır {row_idx}: Bileşen adedi geçersiz")
                    continue
                
                # Create bileşen
                bileşen_id = str(uuid.uuid4())
                bileşen_data = {
                    "id": bileşen_id,
                    "proje_id": proje_id,
                    "proje_adi": proje_adi,
                    "bileşen_adi": bileşen_adi,
                    "malzeme_kodu": malzeme_kodu,
                    "bileşen_adedi": bileşen_adedi,
                    "firma_adi": firma_adi,
                    "iskele_periyodu": "6 Aylık",
                    "gecerlilik_tarihi": gecerlilik_tarihi,
                    "uygunluk": uygunluk,
                    "aciklama": aciklama,
                    "gorseller": [],
                    "created_by": current_user["id"],
                    "created_by_username": current_user.get("username", current_user.get("email", "")),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.iskele_bilesenleri.insert_one(bileşen_data)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Satır {row_idx}: {str(e)}")
                continue
        
        return {
            "message": f"{imported_count} iskele bileşeni başarıyla içe aktarıldı",
            "imported_count": imported_count,
            "errors": errors
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel dosyası işlenemedi: {str(e)}")



@api_router.post("/projeler/bulk-delete")
async def bulk_delete_projeler(proje_ids: List[str], current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    result = await db.projeler.delete_many({"id": {"$in": proje_ids}})
    return {"message": f"{result.deleted_count} proje silindi", "deleted_count": result.deleted_count}

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
