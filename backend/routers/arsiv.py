"""
Arşiv Router - Sistem Arşiv Merkezi
Tüm verileri kategorize edilmiş ZIP dosyası olarak dışa aktarma
"""

from fastapi import APIRouter, HTTPException, Depends, BackgroundTasks
from fastapi.responses import StreamingResponse, JSONResponse
from typing import Optional
import zipfile
import io
import os
import json
from datetime import datetime, timezone
from bson import ObjectId
import asyncio

from database import db
from routers.auth import get_current_user

router = APIRouter(prefix="/arsiv", tags=["Arşiv"])

# Store archive generation progress
archive_progress = {}


class JSONEncoder(json.JSONEncoder):
    """Custom JSON encoder for MongoDB ObjectId and datetime"""
    def default(self, obj):
        if isinstance(obj, ObjectId):
            return str(obj)
        if isinstance(obj, datetime):
            return obj.isoformat()
        return super().default(obj)


def get_upload_path():
    """Get the upload directory path"""
    return os.path.join(os.path.dirname(os.path.dirname(__file__)), "..", "uploads")


@router.get("/progress/{task_id}")
async def get_archive_progress(task_id: str, current_user: dict = Depends(get_current_user)):
    """Get archive generation progress"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    
    if task_id not in archive_progress:
        return {"status": "not_found", "progress": 0, "message": "Görev bulunamadı"}
    
    return archive_progress[task_id]


@router.post("/start")
async def start_archive_generation(background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):
    """Start archive generation in background"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    
    task_id = f"archive_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"
    archive_progress[task_id] = {
        "status": "starting",
        "progress": 0,
        "message": "Arşiv hazırlanıyor...",
        "current_step": "",
        "total_steps": 7
    }
    
    background_tasks.add_task(generate_archive_task, task_id)
    
    return {"task_id": task_id, "message": "Arşiv oluşturma işlemi başlatıldı"}


async def generate_archive_task(task_id: str):
    """Background task for archive generation"""
    try:
        archive_progress[task_id]["status"] = "processing"
        archive_progress[task_id]["progress"] = 5
        archive_progress[task_id]["current_step"] = "Veritabanı bağlantısı kuruluyor..."
        
        # Create in-memory ZIP
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            
            # Step 1: Export Raporlar
            archive_progress[task_id]["progress"] = 10
            archive_progress[task_id]["current_step"] = "Raporlar dışa aktarılıyor..."
            await export_raporlar(zf)
            
            # Step 2: Export İskele Bileşenleri
            archive_progress[task_id]["progress"] = 25
            archive_progress[task_id]["current_step"] = "İskele bileşenleri dışa aktarılıyor..."
            await export_iskele_bilesenleri(zf)
            
            # Step 3: Export Makineler
            archive_progress[task_id]["progress"] = 40
            archive_progress[task_id]["current_step"] = "Makine kayıtları dışa aktarılıyor..."
            await export_makineler(zf)
            
            # Step 4: Export Cephe İskeleleri
            archive_progress[task_id]["progress"] = 55
            archive_progress[task_id]["current_step"] = "Cephe iskeleleri dışa aktarılıyor..."
            await export_cephe_iskeleleri(zf)
            
            # Step 5: Export Raw Database
            archive_progress[task_id]["progress"] = 70
            archive_progress[task_id]["current_step"] = "Veritabanı ham verileri dışa aktarılıyor..."
            await export_raw_database(zf)
            
            # Step 6: Export Media Files
            archive_progress[task_id]["progress"] = 85
            archive_progress[task_id]["current_step"] = "Medya dosyaları ekleniyor..."
            await export_media_files(zf)
            
            # Step 7: Create manifest
            archive_progress[task_id]["progress"] = 95
            archive_progress[task_id]["current_step"] = "Arşiv manifest dosyası oluşturuluyor..."
            await create_manifest(zf)
        
        # Save to temp file
        zip_buffer.seek(0)
        temp_path = f"/tmp/{task_id}.zip"
        with open(temp_path, 'wb') as f:
            f.write(zip_buffer.getvalue())
        
        archive_progress[task_id]["status"] = "completed"
        archive_progress[task_id]["progress"] = 100
        archive_progress[task_id]["current_step"] = "Tamamlandı!"
        archive_progress[task_id]["message"] = "Arşiv başarıyla oluşturuldu"
        archive_progress[task_id]["download_path"] = temp_path
        
    except Exception as e:
        archive_progress[task_id]["status"] = "error"
        archive_progress[task_id]["message"] = f"Hata: {str(e)}"
        archive_progress[task_id]["current_step"] = "Hata oluştu"


async def export_raporlar(zf: zipfile.ZipFile):
    """Export all reports to /Raporlar folder"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    raporlar = await db.raporlar.find({}).to_list(None)
    
    if not raporlar:
        zf.writestr("Raporlar/BOS_KLASOR.txt", "Bu klasörde henüz rapor bulunmamaktadır.")
        return
    
    # Create Excel file for reports
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tüm Raporlar"
    
    # Headers
    headers = ["Rapor No", "Kategori", "Alt Kategori", "Firma", "Proje", "Şehir", 
               "Uygunluk", "Açıklama", "Oluşturulma Tarihi", "Geçerlilik Tarihi"]
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row_idx, rapor in enumerate(raporlar, 2):
        ws.cell(row=row_idx, column=1, value=rapor.get("rapor_no", ""))
        ws.cell(row=row_idx, column=2, value=rapor.get("kategori", ""))
        ws.cell(row=row_idx, column=3, value=rapor.get("alt_kategori", ""))
        ws.cell(row=row_idx, column=4, value=rapor.get("firma_adi", ""))
        ws.cell(row=row_idx, column=5, value=rapor.get("proje_adi", ""))
        ws.cell(row=row_idx, column=6, value=rapor.get("sehir", ""))
        ws.cell(row=row_idx, column=7, value=rapor.get("uygunluk", ""))
        ws.cell(row=row_idx, column=8, value=rapor.get("aciklama", ""))
        ws.cell(row=row_idx, column=9, value=rapor.get("created_at", ""))
        ws.cell(row=row_idx, column=10, value=rapor.get("gecerlilik_tarihi", ""))
    
    # Save Excel to ZIP
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    zf.writestr("Raporlar/tum_raporlar.xlsx", excel_buffer.getvalue())
    
    # Export individual report data with files
    upload_path = get_upload_path()
    for rapor in raporlar:
        rapor_no = rapor.get("rapor_no", "unknown")
        safe_rapor_no = rapor_no.replace("/", "_").replace("\\", "_")
        
        # Create JSON for each report
        rapor_copy = {k: v for k, v in rapor.items() if k != "_id"}
        if "_id" in rapor:
            rapor_copy["_id"] = str(rapor["_id"])
        
        zf.writestr(f"Raporlar/Detay/{safe_rapor_no}.json", 
                   json.dumps(rapor_copy, cls=JSONEncoder, ensure_ascii=False, indent=2))
        
        # Copy associated files
        dosyalar = rapor.get("dosyalar", [])
        for dosya in dosyalar:
            dosya_yolu = dosya.get("dosya_yolu", "")
            if dosya_yolu:
                file_path = os.path.join(upload_path, dosya_yolu.replace("/uploads/", ""))
                if os.path.exists(file_path):
                    with open(file_path, 'rb') as f:
                        filename = os.path.basename(file_path)
                        zf.writestr(f"Raporlar/Dosyalar/{safe_rapor_no}/{filename}", f.read())


async def export_iskele_bilesenleri(zf: zipfile.ZipFile):
    """Export scaffold components to /Iskele_Bilesenleri folder"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    bilesenleri = await db.iskele_bilesenleri.find({}).to_list(None)
    bilesen_adlari = await db.iskele_bilesen_adlari.find({}).to_list(None)
    
    if not bilesenleri and not bilesen_adlari:
        zf.writestr("Iskele_Bilesenleri/BOS_KLASOR.txt", "Bu klasörde henüz bileşen bulunmamaktadır.")
        return
    
    # Create Excel for components
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "İskele Bileşenleri"
    
    headers = ["ID", "Bileşen Adı", "Firma", "Proje", "Miktar", "Durum", "Açıklama", "Oluşturulma Tarihi"]
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, bilesen in enumerate(bilesenleri, 2):
        ws.cell(row=row_idx, column=1, value=bilesen.get("id", ""))
        ws.cell(row=row_idx, column=2, value=bilesen.get("bilesen_adi", ""))
        ws.cell(row=row_idx, column=3, value=bilesen.get("firma_adi", ""))
        ws.cell(row=row_idx, column=4, value=bilesen.get("proje_adi", ""))
        ws.cell(row=row_idx, column=5, value=bilesen.get("miktar", 0))
        ws.cell(row=row_idx, column=6, value=bilesen.get("durum", ""))
        ws.cell(row=row_idx, column=7, value=bilesen.get("aciklama", ""))
        ws.cell(row=row_idx, column=8, value=bilesen.get("created_at", ""))
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    zf.writestr("Iskele_Bilesenleri/tum_bilesenleri.xlsx", excel_buffer.getvalue())
    
    # Export component names
    ws2 = wb.create_sheet("Bileşen Adları")
    ws2.cell(row=1, column=1, value="Bileşen Adı").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Açıklama").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    
    for row_idx, ad in enumerate(bilesen_adlari, 2):
        ws2.cell(row=row_idx, column=1, value=ad.get("bilesen_adi", ""))
        ws2.cell(row=row_idx, column=2, value=ad.get("aciklama", ""))
    
    excel_buffer2 = io.BytesIO()
    wb.save(excel_buffer2)
    excel_buffer2.seek(0)
    zf.writestr("Iskele_Bilesenleri/bilesen_adlari.xlsx", excel_buffer2.getvalue())
    
    # JSON export
    bilesenleri_clean = []
    for b in bilesenleri:
        b_copy = {k: v for k, v in b.items() if k != "_id"}
        if "_id" in b:
            b_copy["_id"] = str(b["_id"])
        bilesenleri_clean.append(b_copy)
    
    zf.writestr("Iskele_Bilesenleri/bilesenleri.json", 
               json.dumps(bilesenleri_clean, cls=JSONEncoder, ensure_ascii=False, indent=2))


async def export_makineler(zf: zipfile.ZipFile):
    """Export machines to /Makine_Takip folder"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    makineler = await db.makineler.find({}).to_list(None)
    operatorler = await db.operatorler.find({}).to_list(None)
    
    # Create README for future structure
    readme_content = """# Makine Takip Arşivi

Bu klasör makine takip verilerini içerir.

## Yapı:
- makineler.xlsx: Tüm makine kayıtları
- operatorler.xlsx: Operatör bilgileri
- makineler.json: Ham makine verileri
- operatorler.json: Ham operatör verileri

## Notlar:
- Teknik belgeler ve sertifikalar ilgili makine klasörlerinde saklanır.
- Bakım geçmişi her makinenin JSON dosyasında yer alır.
"""
    zf.writestr("Makine_Takip/README.md", readme_content)
    
    if makineler:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Makineler"
        
        headers = ["ID", "Makine Adı", "Marka", "Model", "Seri No", "Durum", "Son Bakım", "Açıklama"]
        header_fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, makine in enumerate(makineler, 2):
            ws.cell(row=row_idx, column=1, value=makine.get("id", ""))
            ws.cell(row=row_idx, column=2, value=makine.get("makine_adi", ""))
            ws.cell(row=row_idx, column=3, value=makine.get("marka", ""))
            ws.cell(row=row_idx, column=4, value=makine.get("model", ""))
            ws.cell(row=row_idx, column=5, value=makine.get("seri_no", ""))
            ws.cell(row=row_idx, column=6, value=makine.get("durum", ""))
            ws.cell(row=row_idx, column=7, value=makine.get("son_bakim_tarihi", ""))
            ws.cell(row=row_idx, column=8, value=makine.get("aciklama", ""))
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        zf.writestr("Makine_Takip/makineler.xlsx", excel_buffer.getvalue())
        
        # JSON export
        makineler_clean = []
        for m in makineler:
            m_copy = {k: v for k, v in m.items() if k != "_id"}
            if "_id" in m:
                m_copy["_id"] = str(m["_id"])
            makineler_clean.append(m_copy)
        
        zf.writestr("Makine_Takip/makineler.json",
                   json.dumps(makineler_clean, cls=JSONEncoder, ensure_ascii=False, indent=2))
    
    if operatorler:
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "Operatörler"
        
        headers = ["ID", "Ad Soyad", "TC Kimlik", "Telefon", "Belge Türü", "Belge No", "Durum"]
        for col, header in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=header)
        
        for row_idx, op in enumerate(operatorler, 2):
            ws2.cell(row=row_idx, column=1, value=op.get("id", ""))
            ws2.cell(row=row_idx, column=2, value=op.get("ad_soyad", ""))
            ws2.cell(row=row_idx, column=3, value=op.get("tc_kimlik", ""))
            ws2.cell(row=row_idx, column=4, value=op.get("telefon", ""))
            ws2.cell(row=row_idx, column=5, value=op.get("belge_turu", ""))
            ws2.cell(row=row_idx, column=6, value=op.get("belge_no", ""))
            ws2.cell(row=row_idx, column=7, value=op.get("durum", ""))
        
        excel_buffer2 = io.BytesIO()
        wb2.save(excel_buffer2)
        excel_buffer2.seek(0)
        zf.writestr("Makine_Takip/operatorler.xlsx", excel_buffer2.getvalue())


async def export_cephe_iskeleleri(zf: zipfile.ZipFile):
    """Export facade scaffolding to /Cephe_Iskeleleri folder"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    cephe_iskeleleri = await db.cephe_iskeleleri.find({}).to_list(None)
    projeler = await db.projeler.find({}).to_list(None)
    
    if not cephe_iskeleleri:
        zf.writestr("Cephe_Iskeleleri/BOS_KLASOR.txt", "Bu klasörde henüz cephe iskelesi bulunmamaktadır.")
        return
    
    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cephe İskeleleri"
    
    headers = ["ID", "Proje Adı", "Blok", "Cephe", "Genişlik", "Yükseklik", "Alan", "Durum", "Tarih"]
    header_fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for row_idx, iskele in enumerate(cephe_iskeleleri, 2):
        ws.cell(row=row_idx, column=1, value=iskele.get("id", ""))
        ws.cell(row=row_idx, column=2, value=iskele.get("proje_adi", ""))
        ws.cell(row=row_idx, column=3, value=iskele.get("blok", ""))
        ws.cell(row=row_idx, column=4, value=iskele.get("cephe", ""))
        ws.cell(row=row_idx, column=5, value=iskele.get("genislik", 0))
        ws.cell(row=row_idx, column=6, value=iskele.get("yukseklik", 0))
        ws.cell(row=row_idx, column=7, value=iskele.get("alan", 0))
        ws.cell(row=row_idx, column=8, value=iskele.get("durum", ""))
        ws.cell(row=row_idx, column=9, value=iskele.get("created_at", ""))
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    zf.writestr("Cephe_Iskeleleri/tum_cephe_iskeleleri.xlsx", excel_buffer.getvalue())
    
    # Group by project
    proje_map = {p.get("id"): p.get("proje_adi", "Bilinmeyen") for p in projeler}
    
    proje_iskeleleri = {}
    for iskele in cephe_iskeleleri:
        proje_id = iskele.get("proje_id", "diger")
        proje_adi = proje_map.get(proje_id, iskele.get("proje_adi", "Diger"))
        safe_proje = proje_adi.replace("/", "_").replace("\\", "_")[:50]
        
        if safe_proje not in proje_iskeleleri:
            proje_iskeleleri[safe_proje] = []
        
        iskele_copy = {k: v for k, v in iskele.items() if k != "_id"}
        if "_id" in iskele:
            iskele_copy["_id"] = str(iskele["_id"])
        proje_iskeleleri[safe_proje].append(iskele_copy)
    
    for proje_adi, iskeler in proje_iskeleleri.items():
        zf.writestr(f"Cephe_Iskeleleri/Projeler/{proje_adi}/iskeleleri.json",
                   json.dumps(iskeler, cls=JSONEncoder, ensure_ascii=False, indent=2))


async def export_raw_database(zf: zipfile.ZipFile):
    """Export raw MongoDB collections to /Veritabani_Ham_Veri folder"""
    
    collections = [
        ("users", db.users),
        ("raporlar", db.raporlar),
        ("kategoriler", db.kategoriler),
        ("projeler", db.projeler),
        ("iskele_bilesenleri", db.iskele_bilesenleri),
        ("iskele_bilesen_adlari", db.iskele_bilesen_adlari),
        ("makineler", db.makineler),
        ("operatorler", db.operatorler),
        ("cephe_iskeleleri", db.cephe_iskeleleri),
        ("kalibrasyon_cihazlari", db.kalibrasyon_cihazlari),
        ("notifications", db.notifications),
        ("draws", db.draws),
        ("vocabulary", db.vocabulary),
    ]
    
    summary = {
        "export_date": datetime.now(timezone.utc).isoformat(),
        "collections": {}
    }
    
    for name, collection in collections:
        try:
            docs = await collection.find({}).to_list(None)
            
            # Clean ObjectId
            docs_clean = []
            for doc in docs:
                doc_copy = {}
                for k, v in doc.items():
                    if isinstance(v, ObjectId):
                        doc_copy[k] = str(v)
                    elif isinstance(v, datetime):
                        doc_copy[k] = v.isoformat()
                    else:
                        doc_copy[k] = v
                docs_clean.append(doc_copy)
            
            # Remove sensitive data from users
            if name == "users":
                for doc in docs_clean:
                    if "password" in doc:
                        doc["password"] = "[HIDDEN]"
            
            zf.writestr(f"Veritabani_Ham_Veri/{name}.json",
                       json.dumps(docs_clean, cls=JSONEncoder, ensure_ascii=False, indent=2))
            
            summary["collections"][name] = len(docs_clean)
        except Exception as e:
            summary["collections"][name] = f"Error: {str(e)}"
    
    zf.writestr("Veritabani_Ham_Veri/_summary.json",
               json.dumps(summary, ensure_ascii=False, indent=2))


async def export_media_files(zf: zipfile.ZipFile):
    """Export all media files from uploads directory"""
    upload_path = get_upload_path()
    
    if not os.path.exists(upload_path):
        zf.writestr("Medya_Dosyalari/BOS_KLASOR.txt", "Upload klasörü bulunamadı.")
        return
    
    file_count = 0
    for root, dirs, files in os.walk(upload_path):
        for file in files:
            file_path = os.path.join(root, file)
            relative_path = os.path.relpath(file_path, upload_path)
            
            try:
                with open(file_path, 'rb') as f:
                    zf.writestr(f"Medya_Dosyalari/{relative_path}", f.read())
                    file_count += 1
            except Exception as e:
                continue
    
    if file_count == 0:
        zf.writestr("Medya_Dosyalari/BOS_KLASOR.txt", "Henüz medya dosyası bulunmamaktadır.")


async def create_manifest(zf: zipfile.ZipFile):
    """Create manifest file with archive info"""
    
    # Get counts
    rapor_count = await db.raporlar.count_documents({})
    user_count = await db.users.count_documents({})
    proje_count = await db.projeler.count_documents({})
    kategori_count = await db.kategoriler.count_documents({})
    
    manifest = {
        "archive_info": {
            "created_at": datetime.now(timezone.utc).isoformat(),
            "system": "EKOS - Ekipman Kontrol Otomasyon Sistemi",
            "version": "2.0.0",
            "type": "full_backup"
        },
        "statistics": {
            "total_reports": rapor_count,
            "total_users": user_count,
            "total_projects": proje_count,
            "total_categories": kategori_count
        },
        "folder_structure": {
            "Raporlar": "Tüm muayene raporları ve ekli dosyalar",
            "Iskele_Bilesenleri": "İskele bileşen listesi ve stok bilgileri",
            "Makine_Takip": "Makine kartları ve teknik belgeler",
            "Cephe_Iskeleleri": "Proje bazlı cephe iskele verileri",
            "Veritabani_Ham_Veri": "MongoDB koleksiyonlarının JSON dökümü",
            "Medya_Dosyalari": "Tüm yüklenmiş dosyalar (resim, PDF vb.)"
        },
        "restore_instructions": """
Bu arşivi geri yüklemek için:
1. JSON dosyalarını MongoDB'ye import edin
2. Medya dosyalarını uploads klasörüne kopyalayın
3. Dosya yollarının eşleştiğinden emin olun
"""
    }
    
    zf.writestr("MANIFEST.json", json.dumps(manifest, ensure_ascii=False, indent=2))
    
    # Create human-readable README
    readme = f"""# EKOS Sistem Arşivi

## Oluşturulma Tarihi
{datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M:%S')} UTC

## İstatistikler
- Toplam Rapor: {rapor_count}
- Toplam Kullanıcı: {user_count}
- Toplam Proje: {proje_count}
- Toplam Kategori: {kategori_count}

## Klasör Yapısı
- **Raporlar/** - Tüm muayene raporları ve ekli dosyalar
- **Iskele_Bilesenleri/** - İskele bileşen listesi ve stok bilgileri
- **Makine_Takip/** - Makine kartları ve teknik belgeler
- **Cephe_Iskeleleri/** - Proje bazlı cephe iskele verileri
- **Veritabani_Ham_Veri/** - MongoDB koleksiyonlarının JSON dökümü
- **Medya_Dosyalari/** - Tüm yüklenmiş dosyalar

## Geri Yükleme Talimatları
1. JSON dosyalarını MongoDB'ye import edin
2. Medya dosyalarını uploads klasörüne kopyalayın
3. Dosya yollarının eşleştiğinden emin olun

---
EKOS - Ekipman Kontrol Otomasyon Sistemi v2.0.0
"""
    zf.writestr("README.md", readme)


@router.get("/download/{task_id}")
async def download_archive(task_id: str, current_user: dict = Depends(get_current_user)):
    """Download completed archive"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    
    if task_id not in archive_progress:
        raise HTTPException(status_code=404, detail="Arşiv bulunamadı")
    
    progress = archive_progress[task_id]
    if progress["status"] != "completed":
        raise HTTPException(status_code=400, detail="Arşiv henüz hazır değil")
    
    file_path = progress.get("download_path")
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Arşiv dosyası bulunamadı")
    
    def iterfile():
        with open(file_path, "rb") as f:
            while chunk := f.read(1024 * 1024):  # 1MB chunks
                yield chunk
        # Clean up after download
        try:
            os.remove(file_path)
            del archive_progress[task_id]
        except:
            pass
    
    filename = f"EKOS_Arsiv_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.zip"
    
    return StreamingResponse(
        iterfile(),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
