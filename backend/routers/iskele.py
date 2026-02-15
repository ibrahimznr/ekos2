from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
from pydantic import BaseModel
import io
import uuid

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList

from models import IskeleBileseni, IskeleBileseniCreate
from routers.auth import get_current_user
from database import db

router = APIRouter(tags=["Iskele"])

# Request model for filtered export
class IskeleBileseniFilteredExportRequest(BaseModel):
    firma: Optional[str] = None
    proje_id: Optional[str] = None
    bilesen_adi_search: Optional[str] = None

# ==================== İSKELE BİLEŞEN ADLARI ====================

@router.get("/iskele-bilesen-adlari")
async def get_iskele_bilesen_adlari(current_user: dict = Depends(get_current_user)):
    bilesen_adlari = await db.iskele_bilesen_adlari.find({}, {"_id": 0}).to_list(1000)
    return bilesen_adlari

@router.post("/iskele-bilesen-adlari")
async def create_iskele_bilesen_adi(
    bilesen_adi: str,
    aciklama: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    existing = await db.iskele_bilesen_adlari.find_one({"bilesen_adi": bilesen_adi}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Bu bileşen adı zaten mevcut")
    
    bilesen_id = str(uuid.uuid4())
    bilesen_data = {
        "id": bilesen_id,
        "bilesen_adi": bilesen_adi,
        "aciklama": aciklama,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.iskele_bilesen_adlari.insert_one(bilesen_data)
    created = await db.iskele_bilesen_adlari.find_one({"id": bilesen_id}, {"_id": 0})
    return created

@router.put("/iskele-bilesen-adlari/{bilesen_id}")
async def update_iskele_bilesen_adi(
    bilesen_id: str,
    bilesen_adi: str,
    aciklama: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    existing = await db.iskele_bilesen_adlari.find_one({"id": bilesen_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Bileşen adı bulunamadı")
    
    update_data = {
        "bilesen_adi": bilesen_adi,
        "aciklama": aciklama
    }
    
    await db.iskele_bilesen_adlari.update_one({"id": bilesen_id}, {"$set": update_data})
    updated = await db.iskele_bilesen_adlari.find_one({"id": bilesen_id}, {"_id": 0})
    return updated

@router.delete("/iskele-bilesen-adlari/{bilesen_id}")
async def delete_iskele_bilesen_adi(
    bilesen_id: str,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    result = await db.iskele_bilesen_adlari.delete_one({"id": bilesen_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bileşen adı bulunamadı")
    
    return {"message": "Bileşen adı silindi"}

# ==================== İSKELE BİLEŞENLERİ ====================

@router.get("/iskele-bilesenleri")
async def get_iskele_bilesenleri(
    current_user: dict = Depends(get_current_user),
    limit: int = 500
):
    query = {}
    if current_user.get("role") == "viewer" and current_user.get("firma_adi"):
        query["firma_adi"] = current_user.get("firma_adi")
    
    bilesenleri = await db.iskele_bilesenleri.find(query, {"_id": 0}).to_list(limit)
    return bilesenleri

@router.post("/iskele-bilesenleri")
async def create_iskele_bileseni(
    bilesen: IskeleBileseniCreate,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="İskele bileşeni ekleme yetkiniz yok")
    
    proje = await db.projeler.find_one({"id": bilesen.proje_id}, {"_id": 0})
    if not proje:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    bilesen_id = str(uuid.uuid4())
    bilesen_data = {
        **bilesen.model_dump(),
        "id": bilesen_id,
        "proje_adi": proje.get("proje_adi", ""),
        "iskele_periyodu": "6 Aylık",
        "created_by": current_user["id"],
        "created_by_username": current_user.get("username", current_user.get("email", "")),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.iskele_bilesenleri.insert_one(bilesen_data)
    
    created = await db.iskele_bilesenleri.find_one({"id": bilesen_id}, {"_id": 0})
    return created

@router.get("/iskele-bilesenleri/{bilesen_id}")
async def get_iskele_bileseni(
    bilesen_id: str,
    current_user: dict = Depends(get_current_user)
):
    bilesen = await db.iskele_bilesenleri.find_one({"id": bilesen_id}, {"_id": 0})
    if not bilesen:
        raise HTTPException(status_code=404, detail="İskele bileşeni bulunamadı")
    return bilesen

@router.put("/iskele-bilesenleri/{bilesen_id}")
async def update_iskele_bileseni(
    bilesen_id: str,
    bilesen_update: IskeleBileseniCreate,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="İskele bileşeni güncelleme yetkiniz yok")
    
    existing = await db.iskele_bilesenleri.find_one({"id": bilesen_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="İskele bileşeni bulunamadı")
    
    update_data = bilesen_update.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["iskele_periyodu"] = "6 Aylık"
    
    await db.iskele_bilesenleri.update_one(
        {"id": bilesen_id},
        {"$set": update_data}
    )
    
    updated = await db.iskele_bilesenleri.find_one({"id": bilesen_id}, {"_id": 0})
    return updated

@router.delete("/iskele-bilesenleri/{bilesen_id}")
async def delete_iskele_bileseni(
    bilesen_id: str,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="İskele bileşeni silme yetkiniz yok")
    
    result = await db.iskele_bilesenleri.delete_one({"id": bilesen_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="İskele bileşeni bulunamadı")
    
    return {"message": "İskele bileşeni silindi"}

@router.post("/iskele-bilesenleri/bulk-delete")
async def bulk_delete_iskele_bilesenleri_route(
    bilesen_ids: List[str],
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="İskele bileşeni silme yetkiniz yok")
    
    if not bilesen_ids:
        raise HTTPException(status_code=400, detail="Silinecek bileşen ID'leri belirtilmedi")
    
    result = await db.iskele_bilesenleri.delete_many({"id": {"$in": bilesen_ids}})
    return {"message": f"{result.deleted_count} iskele bileşeni silindi", "deleted_count": result.deleted_count}

# ==================== İSKELE EXCEL ====================

@router.get("/iskele-bilesenleri/excel/export")
async def export_iskele_excel(current_user: dict = Depends(get_current_user)):
    query = {}
    if current_user.get("role") == "viewer" and current_user.get("firma_adi"):
        query["firma_adi"] = current_user.get("firma_adi")
    
    bilesenleri = await db.iskele_bilesenleri.find(query, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "İskele Bileşenleri"
    
    headers = [
        "Bileşen Adı", "Malzeme Kodu", "Bileşen Adedi", "Firma Adı",
        "Geçerlilik Tarihi", "Uygunluk", "Açıklama", "Proje Adı"
    ]
    
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, bilesen in enumerate(bilesenleri, 2):
        ws.cell(row=row_idx, column=1, value=bilesen.get("bileşen_adi", ""))
        ws.cell(row=row_idx, column=2, value=bilesen.get("malzeme_kodu", ""))
        ws.cell(row=row_idx, column=3, value=bilesen.get("bileşen_adedi", 0))
        ws.cell(row=row_idx, column=4, value=bilesen.get("firma_adi", ""))
        ws.cell(row=row_idx, column=5, value=bilesen.get("gecerlilik_tarihi", ""))
        ws.cell(row=row_idx, column=6, value=bilesen.get("uygunluk", ""))
        ws.cell(row=row_idx, column=7, value=bilesen.get("aciklama", ""))
        ws.cell(row=row_idx, column=8, value=bilesen.get("proje_adi", ""))
    
    for col in ws.columns:
        column = col[0].column_letter
        ws.column_dimensions[column].width = 20
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=iskele_bilesenleri_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@router.get("/iskele-bilesenleri/excel/template")
async def download_iskele_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "İskele Bileşenleri Şablonu"
    
    headers = [
        "Bileşen Adı", "Malzeme Kodu", "Bileşen Adedi", "Firma Adı",
        "Geçerlilik Tarihi", "Uygunluk", "Açıklama"
    ]
    
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    examples = [
        ["Çelik Direk", "CD-001", 10, "ABC İnşaat", "2025-12-31", "Uygun", "Standart çelik direk"],
        ["Bağlantı Elemanı", "BE-002", 50, "XYZ Yapı", "2025-06-30", "Uygun", ""],
        ["Destek Parçası", "DP-003", 25, "Test Firma", "", "Uygun Değil", "Kontrol gerekli"]
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col, value in enumerate(example, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    for col in ws.columns:
        column = col[0].column_letter
        ws.column_dimensions[column].width = 20
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=iskele_bilesenleri_sablonu.xlsx"}
    )

@router.post("/iskele-bilesenleri/excel/import")
async def import_iskele_excel(
    file: UploadFile = File(...),
    proje_id: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="İskele bileşeni içe aktarma yetkiniz yok")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sadece Excel dosyaları yüklenebilir")
    
    proje = await db.projeler.find_one({"id": proje_id}, {"_id": 0})
    if not proje:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    proje_adi = proje.get("proje_adi", "")
    
    content = await file.read()
    excel_file = io.BytesIO(content)
    
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        imported_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):
                continue
            
            try:
                def get_cell(index):
                    return row[index] if index < len(row) and row[index] is not None else None
                
                bilesen_adi = str(get_cell(0)) if get_cell(0) else ""
                malzeme_kodu = str(get_cell(1)) if get_cell(1) else ""
                bilesen_adedi_raw = get_cell(2)
                firma_adi = str(get_cell(3)) if get_cell(3) else ""
                gecerlilik_tarihi = str(get_cell(4)) if get_cell(4) else None
                uygunluk = str(get_cell(5)) if get_cell(5) else "Uygun"
                aciklama = str(get_cell(6)) if get_cell(6) else None
                
                if not bilesen_adi or not malzeme_kodu or not firma_adi:
                    errors.append(f"Satır {row_idx}: Zorunlu alanlar eksik")
                    continue
                
                try:
                    bilesen_adedi = int(bilesen_adedi_raw) if bilesen_adedi_raw else 1
                    if bilesen_adedi < 1:
                        errors.append(f"Satır {row_idx}: Bileşen adedi en az 1 olmalıdır")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Satır {row_idx}: Bileşen adedi geçersiz")
                    continue
                
                bilesen_id = str(uuid.uuid4())
                bilesen_data = {
                    "id": bilesen_id,
                    "proje_id": proje_id,
                    "proje_adi": proje_adi,
                    "bileşen_adi": bilesen_adi,
                    "malzeme_kodu": malzeme_kodu,
                    "bileşen_adedi": bilesen_adedi,
                    "firma_adi": firma_adi,
                    "iskele_periyodu": "6 Aylık",
                    "gecerlilik_tarihi": gecerlilik_tarihi,
                    "uygunluk": uygunluk,
                    "aciklama": aciklama,
                    "gorseller": [],
                    "created_by": current_user["id"],
                    "created_by_username": current_user.get("username", current_user.get("email", "")),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.iskele_bilesenleri.insert_one(bilesen_data)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Satır {row_idx}: {str(e)}")
                continue
        
        return {
            "message": f"{imported_count} iskele bileşeni başarıyla içe aktarıldı",
            "imported_count": imported_count,
            "errors": errors
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel dosyası işlenemedi: {str(e)}")


# ==================== FİLTRELENMİŞ İSTATİSTİKLER VE EXCEL EXPORT ====================

@router.get("/iskele-bilesenleri/stats/filtered")
async def get_iskele_bilesenleri_filtered_stats(
    firma: Optional[str] = Query(None),
    proje_id: Optional[str] = Query(None),
    bilesen_adi_search: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user)
):
    """Filtrelenmiş iskele bileşenleri istatistiklerini getir"""
    query = {}
    
    if firma and firma != 'all':
        query["firma_adi"] = firma
    if proje_id and proje_id != 'all':
        query["proje_id"] = proje_id
    
    bilesenleri = await db.iskele_bilesenleri.find(query, {"_id": 0}).to_list(100000)
    
    # Apply search filter in Python (for partial match)
    if bilesen_adi_search:
        search_lower = bilesen_adi_search.lower()
        bilesenleri = [b for b in bilesenleri if search_lower in (b.get('bileşen_adi', '') or '').lower()]
    
    total = len(bilesenleri)
    uygun = sum(1 for b in bilesenleri if b.get('uygunluk') == 'Uygun')
    uygun_degil = sum(1 for b in bilesenleri if b.get('uygunluk') == 'Uygun Değil')
    uygunluk_orani = round((uygun / total) * 100, 1) if total > 0 else 0
    
    # Bileşen dağılımı
    bilesen_map = {}
    for b in bilesenleri:
        adi = b.get('bileşen_adi')
        if adi:
            bilesen_map[adi] = bilesen_map.get(adi, 0) + b.get('bileşen_adedi', 1)
    
    bilesen_dagilim = [
        {"bileşen_adi": k, "count": v}
        for k, v in sorted(bilesen_map.items(), key=lambda x: x[1], reverse=True)
    ]
    
    return {
        "total": total,
        "uygun": uygun,
        "uygun_degil": uygun_degil,
        "uygunluk_orani": uygunluk_orani,
        "bilesen_dagilim": bilesen_dagilim
    }


@router.get("/iskele-bilesenleri/filter-options")
async def get_iskele_filter_options(current_user: dict = Depends(get_current_user)):
    """Filtreleme için firma ve proje listesini getir"""
    bilesenleri = await db.iskele_bilesenleri.find({}, {"_id": 0, "firma_adi": 1, "proje_id": 1, "proje_adi": 1}).to_list(100000)
    
    firmalar = set()
    projeler = {}
    
    for b in bilesenleri:
        if b.get('firma_adi'):
            firmalar.add(b['firma_adi'])
        if b.get('proje_id') and b.get('proje_adi'):
            projeler[b['proje_id']] = b['proje_adi']
    
    return {
        "firmalar": sorted(list(firmalar)),
        "projeler": [{"id": k, "adi": v} for k, v in sorted(projeler.items(), key=lambda x: x[1])]
    }


@router.post("/iskele-bilesenleri/excel/export-filtered")
async def export_iskele_bilesenleri_filtered(
    request: IskeleBileseniFilteredExportRequest,
    current_user: dict = Depends(get_current_user)
):
    """Filtrelenmiş iskele bileşenlerini Dashboard görünümünde Excel'e aktar"""
    query = {}
    
    if request.firma and request.firma != 'all':
        query["firma_adi"] = request.firma
    if request.proje_id and request.proje_id != 'all':
        query["proje_id"] = request.proje_id
    
    bilesenleri = await db.iskele_bilesenleri.find(query, {"_id": 0}).to_list(100000)
    
    # Apply search filter
    if request.bilesen_adi_search:
        search_lower = request.bilesen_adi_search.lower()
        bilesenleri = [b for b in bilesenleri if search_lower in (b.get('bileşen_adi', '') or '').lower()]
    
    if not bilesenleri:
        raise HTTPException(status_code=404, detail="Filtrelere uyan bileşen bulunamadı")
    
    # Calculate stats
    today = datetime.now(timezone.utc)
    total = len(bilesenleri)
    uygun = sum(1 for b in bilesenleri if b.get('uygunluk') == 'Uygun')
    uygun_degil = sum(1 for b in bilesenleri if b.get('uygunluk') == 'Uygun Değil')
    uygunluk_orani = round((uygun / total) * 100, 1) if total > 0 else 0
    
    # Bileşen dağılımı
    bilesen_map = {}
    for b in bilesenleri:
        adi = b.get('bileşen_adi')
        if adi:
            bilesen_map[adi] = bilesen_map.get(adi, 0) + b.get('bileşen_adedi', 1)
    bilesen_list = sorted(bilesen_map.items(), key=lambda x: x[1], reverse=True)
    
    # Firma dağılımı
    firma_map = {}
    for b in bilesenleri:
        firma = b.get('firma_adi')
        if firma:
            firma_map[firma] = firma_map.get(firma, 0) + b.get('bileşen_adedi', 1)
    firma_list = sorted(firma_map.items(), key=lambda x: x[1], reverse=True)
    
    # Create workbook
    wb = Workbook()
    
    # ===== DASHBOARD SHEET =====
    ws_dashboard = wb.active
    ws_dashboard.title = "Dashboard"
    ws_dashboard.sheet_view.showGridLines = False
    
    # Set column widths
    for col in range(1, 25):
        ws_dashboard.column_dimensions[get_column_letter(col)].width = 6
    
    # Colors
    teal_border = Side(style='medium', color='0d9488')
    green_border = Side(style='medium', color='16a34a')
    red_border = Side(style='medium', color='dc2626')
    blue_border = Side(style='medium', color='2563eb')
    
    teal_fill = PatternFill(start_color='ccfbf1', end_color='ccfbf1', fill_type='solid')
    green_fill = PatternFill(start_color='dcfce7', end_color='dcfce7', fill_type='solid')
    red_fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')
    blue_fill = PatternFill(start_color='dbeafe', end_color='dbeafe', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    header_fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
    
    # White background
    for row in range(1, 50):
        for col in range(1, 25):
            ws_dashboard.cell(row=row, column=col).fill = white_fill
    
    # Title
    title_cell = ws_dashboard.cell(row=2, column=2, value="İskele Bileşenleri Dashboard")
    title_cell.font = Font(bold=True, size=18, color='0d9488')
    ws_dashboard.merge_cells('B2:T2')
    
    # Filter info
    filter_info = []
    if request.firma and request.firma != 'all':
        filter_info.append(f"Firma: {request.firma}")
    if request.proje_id and request.proje_id != 'all':
        proje = await db.projeler.find_one({"id": request.proje_id}, {"_id": 0})
        if proje:
            filter_info.append(f"Proje: {proje.get('proje_adi', '')}")
    if request.bilesen_adi_search:
        filter_info.append(f"Arama: {request.bilesen_adi_search}")
    
    filter_text = " | ".join(filter_info) if filter_info else "Tüm Veriler"
    filter_cell = ws_dashboard.cell(row=3, column=2, value=f"Filtreler: {filter_text}  |  Oluşturma: {today.strftime('%d.%m.%Y %H:%M')}")
    filter_cell.font = Font(size=10, color='6b7280', italic=True)
    ws_dashboard.merge_cells('B3:T3')
    
    # Helper function to create card
    def create_card(start_row, start_col, end_col, title, value, border_color, fill_color, text_color):
        title_cell = ws_dashboard.cell(row=start_row, column=start_col, value=title)
        title_cell.font = Font(bold=True, size=10, color='374151')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        value_cell = ws_dashboard.cell(row=start_row + 1, column=start_col, value=value)
        value_cell.font = Font(bold=True, size=28, color=text_color)
        value_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws_dashboard.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
        ws_dashboard.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 1, end_column=end_col)
        
        for row in range(start_row, start_row + 2):
            for col in range(start_col, end_col + 1):
                cell = ws_dashboard.cell(row=row, column=col)
                cell.fill = fill_color
                left = border_color if col == start_col else None
                right = border_color if col == end_col else None
                top = border_color if row == start_row else None
                bottom = border_color if row == start_row + 1 else None
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)
    
    # Create stat cards (Row 5-6)
    create_card(5, 2, 5, "Toplam", total, teal_border, teal_fill, '0d9488')
    create_card(5, 7, 10, "Uygun", uygun, green_border, green_fill, '16a34a')
    create_card(5, 12, 15, "Uygun Değil", uygun_degil, red_border, red_fill, 'dc2626')
    create_card(5, 17, 20, f"Oran (%{uygunluk_orani})", f"{uygunluk_orani}%", blue_border, blue_fill, '2563eb')
    
    # Distribution tables
    row_colors = ['ccfbf1', 'a7f3d0', 'fef3c7', 'fecaca', 'ddd6fe', 'bfdbfe', 'fbcfe8', 'e0e7ff']
    
    def create_distribution_table(start_row, start_col, title, data_list, max_items=10):
        t_cell = ws_dashboard.cell(row=start_row, column=start_col, value=title)
        t_cell.font = Font(bold=True, size=12, color='1f2937')
        ws_dashboard.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 5)
        
        ws_dashboard.cell(row=start_row + 2, column=start_col, value="Ad").font = Font(bold=True, size=10, color='374151')
        ws_dashboard.cell(row=start_row + 2, column=start_col).fill = header_fill
        ws_dashboard.merge_cells(start_row=start_row + 2, start_column=start_col, end_row=start_row + 2, end_column=start_col + 3)
        
        ws_dashboard.cell(row=start_row + 2, column=start_col + 4, value="Adet").font = Font(bold=True, size=10, color='374151')
        ws_dashboard.cell(row=start_row + 2, column=start_col + 4).fill = header_fill
        ws_dashboard.cell(row=start_row + 2, column=start_col + 4).alignment = Alignment(horizontal='center')
        
        ws_dashboard.cell(row=start_row + 2, column=start_col + 5, value="Oran").font = Font(bold=True, size=10, color='374151')
        ws_dashboard.cell(row=start_row + 2, column=start_col + 5).fill = header_fill
        ws_dashboard.cell(row=start_row + 2, column=start_col + 5).alignment = Alignment(horizontal='center')
        
        for idx, (name, count) in enumerate(data_list[:max_items]):
            row_num = start_row + 3 + idx
            percentage = round((count / total) * 100) if total > 0 else 0
            color = row_colors[idx % len(row_colors)]
            row_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            
            name_cell = ws_dashboard.cell(row=row_num, column=start_col, value=name[:25] if len(name) > 25 else name)
            name_cell.font = Font(size=9, color='374151')
            name_cell.fill = row_fill
            ws_dashboard.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=start_col + 3)
            
            count_cell = ws_dashboard.cell(row=row_num, column=start_col + 4, value=count)
            count_cell.font = Font(size=9, color='374151', bold=True)
            count_cell.fill = row_fill
            count_cell.alignment = Alignment(horizontal='center')
            
            pct_cell = ws_dashboard.cell(row=row_num, column=start_col + 5, value=f"%{percentage}")
            pct_cell.font = Font(size=9, color='6b7280')
            pct_cell.fill = row_fill
            pct_cell.alignment = Alignment(horizontal='center')
    
    # Bileşen Dağılımı (Row 9, Col B)
    create_distribution_table(9, 2, "Bileşen Dağılımı", bilesen_list, 10)
    
    # Firma Dağılımı (Row 9, Col I)
    create_distribution_table(9, 9, "Firma Dağılımı", firma_list, 10)
    
    # Pie chart data (hidden sheet)
    ws_chart_data = wb.create_sheet("ChartData")
    ws_chart_data.cell(row=1, column=1, value="Uygunluk")
    ws_chart_data.cell(row=1, column=2, value="Adet")
    ws_chart_data.cell(row=2, column=1, value="Uygun")
    ws_chart_data.cell(row=2, column=2, value=uygun)
    ws_chart_data.cell(row=3, column=1, value="Uygun Değil")
    ws_chart_data.cell(row=3, column=2, value=uygun_degil)
    ws_chart_data.sheet_state = 'hidden'
    
    # Pie chart
    pie = PieChart()
    pie.title = "Uygunluk Durumu"
    labels = Reference(ws_chart_data, min_col=1, min_row=2, max_row=3)
    data = Reference(ws_chart_data, min_col=2, min_row=1, max_row=3)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    pie.width = 10
    pie.height = 7
    ws_dashboard.add_chart(pie, "P9")
    
    # ===== DATA SHEET =====
    ws_data = wb.create_sheet("Bileşenler")
    
    headers = ["Bileşen Adı", "Malzeme Kodu", "Adet", "Firma", "Proje", "Uygunluk", "Geçerlilik", "Açıklama"]
    header_fill_data = PatternFill(start_color="0d9488", end_color="0d9488", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.fill = header_fill_data
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, bilesen in enumerate(bilesenleri, 2):
        ws_data.cell(row=row_idx, column=1, value=bilesen.get("bileşen_adi", ""))
        ws_data.cell(row=row_idx, column=2, value=bilesen.get("malzeme_kodu", ""))
        ws_data.cell(row=row_idx, column=3, value=bilesen.get("bileşen_adedi", 1))
        ws_data.cell(row=row_idx, column=4, value=bilesen.get("firma_adi", ""))
        ws_data.cell(row=row_idx, column=5, value=bilesen.get("proje_adi", ""))
        ws_data.cell(row=row_idx, column=6, value=bilesen.get("uygunluk", ""))
        ws_data.cell(row=row_idx, column=7, value=bilesen.get("gecerlilik_tarihi", ""))
        ws_data.cell(row=row_idx, column=8, value=bilesen.get("aciklama", ""))
    
    # Auto-fit columns
    for col in ws_data.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_data.column_dimensions[column].width = min(max_length + 2, 40)
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Filename
    filter_parts = []
    if request.firma and request.firma != 'all':
        filter_parts.append(request.firma[:10])
    if request.proje_id and request.proje_id != 'all':
        filter_parts.append("proje")
    if request.bilesen_adi_search:
        filter_parts.append("arama")
    
    filter_suffix = "_".join(filter_parts) if filter_parts else "tum"
    filename = f"iskele_bilesenleri_{filter_suffix}_{total}_adet.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
