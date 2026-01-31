from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List
from datetime import datetime, timezone
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from models import Kategori, KategoriCreate
from routers.auth import get_current_user
from database import db

router = APIRouter(prefix="/kategoriler", tags=["Kategoriler"])

@router.get("", response_model=List[Kategori])
async def get_kategoriler(current_user: dict = Depends(get_current_user)):
    kategoriler = await db.kategoriler.find({}, {"_id": 0}).to_list(1000)
    for kat in kategoriler:
        if isinstance(kat['created_at'], str):
            kat['created_at'] = datetime.fromisoformat(kat['created_at'])
    return kategoriler

@router.get("/excel/template")
async def download_kategori_template(current_user: dict = Depends(get_current_user)):
    """Kategori import şablonunu indir"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Kategoriler"
    
    # Headers
    headers = ["Kategori Adı*", "Alt Kategoriler (virgülle ayırın)", "Açıklama"]
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Example rows
    examples = [
        ("Kaldırma-İletme", "Vinç, Forklift, Transpalet, Platform", "Kaldırma ve iletme ekipmanları"),
        ("Basınçlı Kaplar", "Kompresör, Hava Tankı, LPG Tankı", "Basınçlı kap muayeneleri"),
        ("Elektrik", "Pano, Topraklama, Paratoner", "Elektrik tesisatı kontrolleri"),
    ]
    
    for row_idx, (isim, alt_kat, aciklama) in enumerate(examples, 2):
        ws.cell(row=row_idx, column=1, value=isim)
        ws.cell(row=row_idx, column=2, value=alt_kat)
        ws.cell(row=row_idx, column=3, value=aciklama)
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 40
    
    # Instructions sheet
    ws_info = wb.create_sheet("Bilgi")
    ws_info.cell(row=1, column=1, value="EKOS Kategori Import Şablonu").font = Font(bold=True, size=14)
    ws_info.cell(row=3, column=1, value="Kullanım Talimatları:")
    ws_info.cell(row=4, column=1, value="1. 'Kategoriler' sayfasına verilerinizi girin")
    ws_info.cell(row=5, column=1, value="2. 'Kategori Adı' alanı zorunludur (*)")
    ws_info.cell(row=6, column=1, value="3. Alt kategorileri virgül (,) ile ayırarak yazın")
    ws_info.cell(row=7, column=1, value="4. Örnek satırları silebilir veya üzerine yazabilirsiniz")
    ws_info.cell(row=8, column=1, value="5. Mevcut kategoriler varsa güncellenmez, sadece yeniler eklenir")
    ws_info.column_dimensions['A'].width = 60
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=kategori_sablonu.xlsx"}
    )

@router.post("/excel/import")
async def import_kategoriler_excel(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Excel dosyasından kategorileri toplu import et"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sadece Excel dosyaları (.xlsx, .xls) kabul edilir")
    
    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Get existing categories to avoid duplicates
        existing_kategoriler = await db.kategoriler.find({}, {"_id": 0, "isim": 1}).to_list(1000)
        existing_names = {k['isim'].lower().strip() for k in existing_kategoriler}
        
        imported_count = 0
        skipped_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:  # Skip empty rows
                continue
                
            isim = str(row[0]).strip() if row[0] else None
            
            if not isim:
                continue
            
            # Check for duplicate
            if isim.lower() in existing_names:
                skipped_count += 1
                continue
            
            # Parse alt kategoriler
            alt_kategoriler = []
            if len(row) > 1 and row[1]:
                alt_kat_str = str(row[1]).strip()
                alt_kategoriler = [ak.strip() for ak in alt_kat_str.split(',') if ak.strip()]
            
            # Parse açıklama
            aciklama = str(row[2]).strip() if len(row) > 2 and row[2] else None
            
            try:
                kategori = Kategori(
                    isim=isim,
                    alt_kategoriler=alt_kategoriler,
                    aciklama=aciklama
                )
                doc = kategori.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                await db.kategoriler.insert_one(doc)
                existing_names.add(isim.lower())
                imported_count += 1
            except Exception as e:
                errors.append(f"Satır {row_idx}: {str(e)}")
        
        result_message = f"{imported_count} kategori başarıyla eklendi"
        if skipped_count > 0:
            result_message += f", {skipped_count} kategori zaten mevcut (atlandı)"
        if errors:
            result_message += f", {len(errors)} hata oluştu"
        
        return {
            "message": result_message,
            "imported_count": imported_count,
            "skipped_count": skipped_count,
            "errors": errors[:10]  # İlk 10 hatayı döndür
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel dosyası işlenirken hata: {str(e)}")

@router.post("", response_model=Kategori)
async def create_kategori(kategori_create: KategoriCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    kategori = Kategori(**kategori_create.model_dump())
    doc = kategori.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.kategoriler.insert_one(doc)
    return kategori

@router.put("/{kategori_id}")
async def update_kategori(kategori_id: str, kategori_update: KategoriCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    existing = await db.kategoriler.find_one({"id": kategori_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Kategori bulunamadı")
    
    update_data = kategori_update.model_dump()
    await db.kategoriler.update_one({"id": kategori_id}, {"$set": update_data})
    
    updated_kategori = await db.kategoriler.find_one({"id": kategori_id}, {"_id": 0})
    return updated_kategori

@router.delete("/{kategori_id}")
async def delete_kategori(kategori_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    result = await db.kategoriler.delete_one({"id": kategori_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kategori bulunamadı")
    return {"message": "Kategori silindi"}

@router.post("/bulk-delete")
async def bulk_delete_kategoriler(kategori_ids: List[str], current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    result = await db.kategoriler.delete_many({"id": {"$in": kategori_ids}})
    return {"message": f"{result.deleted_count} kategori silindi", "deleted_count": result.deleted_count}
