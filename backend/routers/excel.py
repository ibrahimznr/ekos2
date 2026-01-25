from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
from pathlib import Path
from pydantic import BaseModel
import io
import uuid

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import Rapor
from routers.auth import get_current_user
from database import db
from utils import generate_rapor_no
from constants import SEHIRLER

router = APIRouter(prefix="/excel", tags=["Excel"])

# Request model for selective export
class ExcelExportRequest(BaseModel):
    rapor_ids: List[str]

# Request model for filtered export
class FilteredExcelExportRequest(BaseModel):
    proje_id: Optional[str] = None
    sehir: Optional[str] = None
    firma: Optional[str] = None

@router.post("/export")
async def export_excel_selected(request: ExcelExportRequest, current_user: dict = Depends(get_current_user)):
    """Seçili raporları Excel'e aktar"""
    if not request.rapor_ids:
        raise HTTPException(status_code=400, detail="En az bir rapor seçilmelidir")
    
    raporlar = await db.raporlar.find({"id": {"$in": request.rapor_ids}}, {"_id": 0}).to_list(10000)
    
    if not raporlar:
        raise HTTPException(status_code=404, detail="Seçilen raporlar bulunamadı")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Raporlar"
    
    headers = [
        "Rapor No", "Ekipman Adı", "Kategori", "Firma", "Lokasyon",
        "Marka/Model", "Seri No", "Alt Kategori", "Periyot",
        "Geçerlilik Tarihi", "Uygunluk", "Proje", "Şehir", "Açıklama", "Oluşturma Tarihi"
    ]
    
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, rapor in enumerate(raporlar, 2):
        ws.cell(row=row_idx, column=1, value=rapor.get("rapor_no", ""))
        ws.cell(row=row_idx, column=2, value=rapor.get("ekipman_adi", ""))
        ws.cell(row=row_idx, column=3, value=rapor.get("kategori", ""))
        ws.cell(row=row_idx, column=4, value=rapor.get("firma", ""))
        ws.cell(row=row_idx, column=5, value=rapor.get("lokasyon", ""))
        ws.cell(row=row_idx, column=6, value=rapor.get("marka_model", ""))
        ws.cell(row=row_idx, column=7, value=rapor.get("seri_no", ""))
        ws.cell(row=row_idx, column=8, value=rapor.get("alt_kategori", ""))
        ws.cell(row=row_idx, column=9, value=rapor.get("periyot", ""))
        ws.cell(row=row_idx, column=10, value=rapor.get("gecerlilik_tarihi", ""))
        ws.cell(row=row_idx, column=11, value=rapor.get("uygunluk", ""))
        ws.cell(row=row_idx, column=12, value=rapor.get("proje_adi", ""))
        ws.cell(row=row_idx, column=13, value=rapor.get("sehir", ""))
        ws.cell(row=row_idx, column=14, value=rapor.get("aciklama", ""))
        created_at = rapor.get("created_at", "")
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        ws.cell(row=row_idx, column=15, value=created_at.strftime("%Y-%m-%d %H:%M") if created_at else "")
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=raporlar_{len(raporlar)}_adet.xlsx"}
    )

@router.post("/export-filtered")
async def export_excel_filtered(request: FilteredExcelExportRequest, current_user: dict = Depends(get_current_user)):
    """Filtrelenmiş raporları Dashboard görünümünde Excel'e aktar"""
    
    # Build query based on filters
    query = {}
    if request.proje_id and request.proje_id != 'all':
        query["proje_id"] = request.proje_id
    if request.sehir and request.sehir != 'all':
        query["sehir"] = request.sehir
    if request.firma and request.firma != 'all':
        query["firma"] = request.firma
    
    raporlar = await db.raporlar.find(query, {"_id": 0}).to_list(10000)
    
    if not raporlar:
        raise HTTPException(status_code=404, detail="Filtrelere uyan rapor bulunamadı")
    
    # Calculate statistics
    today = datetime.now(timezone.utc)
    current_month = today.month
    current_year = today.year
    
    total_count = len(raporlar)
    uygun_count = len([r for r in raporlar if r.get('uygunluk') == 'Uygun'])
    uygun_degil_count = len([r for r in raporlar if r.get('uygunluk') == 'Uygun Değil'])
    
    # This month's reports
    monthly_count = 0
    for r in raporlar:
        created_at = r.get('created_at')
        if created_at:
            if isinstance(created_at, str):
                try:
                    created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                except:
                    continue
            if created_at.month == current_month and created_at.year == current_year:
                monthly_count += 1
    
    # Category distribution
    kategori_map = {}
    for r in raporlar:
        kategori = r.get('kategori')
        if kategori:
            kategori_map[kategori] = kategori_map.get(kategori, 0) + 1
    
    kategori_list = sorted(kategori_map.items(), key=lambda x: x[1], reverse=True)
    
    # Create workbook
    wb = Workbook()
    
    # ===== DASHBOARD SHEET =====
    ws_dashboard = wb.active
    ws_dashboard.title = "Dashboard"
    
    # Hide gridlines
    ws_dashboard.sheet_view.showGridLines = False
    
    # Set column widths for card layout
    for col in range(1, 20):
        ws_dashboard.column_dimensions[get_column_letter(col)].width = 5
    
    # Row heights
    ws_dashboard.row_dimensions[1].height = 20
    ws_dashboard.row_dimensions[2].height = 25
    ws_dashboard.row_dimensions[3].height = 15
    ws_dashboard.row_dimensions[4].height = 50
    ws_dashboard.row_dimensions[5].height = 20
    ws_dashboard.row_dimensions[6].height = 20
    
    # Define colors
    blue_border = Side(style='medium', color='1e40af')
    purple_border = Side(style='medium', color='7c3aed')
    green_border = Side(style='medium', color='16a34a')
    red_border = Side(style='medium', color='dc2626')
    
    blue_fill = PatternFill(start_color='dbeafe', end_color='dbeafe', fill_type='solid')
    purple_fill = PatternFill(start_color='ede9fe', end_color='ede9fe', fill_type='solid')
    green_fill = PatternFill(start_color='dcfce7', end_color='dcfce7', fill_type='solid')
    red_fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')
    
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Apply white background to entire visible area
    for row in range(1, 35):
        for col in range(1, 20):
            ws_dashboard.cell(row=row, column=col).fill = white_fill
    
    # Title
    title_cell = ws_dashboard.cell(row=2, column=2, value="EKOS Dashboard Raporu")
    title_cell.font = Font(bold=True, size=18, color='1e40af')
    ws_dashboard.merge_cells('B2:R2')
    
    # Filter info
    filter_info = []
    if request.proje_id and request.proje_id != 'all':
        proje = await db.projeler.find_one({"id": request.proje_id}, {"_id": 0})
        if proje:
            filter_info.append(f"Proje: {proje.get('proje_adi', '')}")
    if request.sehir and request.sehir != 'all':
        filter_info.append(f"İl: {request.sehir}")
    if request.firma and request.firma != 'all':
        filter_info.append(f"Firma: {request.firma}")
    
    filter_text = " | ".join(filter_info) if filter_info else "Tüm Veriler"
    filter_cell = ws_dashboard.cell(row=3, column=2, value=f"Filtreler: {filter_text}  |  Oluşturma: {today.strftime('%d.%m.%Y %H:%M')}")
    filter_cell.font = Font(size=10, color='6b7280', italic=True)
    ws_dashboard.merge_cells('B3:R3')
    
    # Helper function to create card
    def create_card(start_row, start_col, end_col, title, value, border_color, fill_color, text_color):
        # Title row
        title_cell = ws_dashboard.cell(row=start_row, column=start_col, value=title)
        title_cell.font = Font(bold=True, size=10, color='374151')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Value row  
        value_cell = ws_dashboard.cell(row=start_row + 1, column=start_col, value=value)
        value_cell.font = Font(bold=True, size=28, color=text_color)
        value_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells for title and value
        ws_dashboard.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
        ws_dashboard.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 1, end_column=end_col)
        
        # Apply fill and border to all cells in card
        border = Border(left=border_color, right=border_color, top=border_color, bottom=border_color)
        for row in range(start_row, start_row + 2):
            for col in range(start_col, end_col + 1):
                cell = ws_dashboard.cell(row=row, column=col)
                cell.fill = fill_color
                # Apply border only to edge cells
                left = border_color if col == start_col else None
                right = border_color if col == end_col else None
                top = border_color if row == start_row else None
                bottom = border_color if row == start_row + 1 else None
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)
    
    # Create stat cards (Row 5-6)
    # Card 1: Toplam Rapor (Blue)
    create_card(5, 2, 5, "Toplam Rapor", total_count, blue_border, blue_fill, '1e40af')
    
    # Card 2: Bu Ay (Purple)
    create_card(5, 7, 10, "Bu Ay", monthly_count, purple_border, purple_fill, '7c3aed')
    
    # Card 3: Uygun (Green)
    create_card(5, 12, 15, "Uygun", uygun_count, green_border, green_fill, '16a34a')
    
    # Card 4: Uygun Değil (Red)
    create_card(5, 17, 20, "Uygun Değil", uygun_degil_count, red_border, red_fill, 'dc2626')
    
    # Kategori Dağılımı section
    kategori_title = ws_dashboard.cell(row=9, column=2, value="Kategori Dağılımı")
    kategori_title.font = Font(bold=True, size=14, color='1f2937')
    ws_dashboard.merge_cells('B9:H9')
    
    # Kategori table header
    header_fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
    header_border = Border(
        bottom=Side(style='thin', color='d1d5db')
    )
    
    ws_dashboard.cell(row=11, column=2, value="Kategori").font = Font(bold=True, size=11, color='374151')
    ws_dashboard.cell(row=11, column=2).fill = header_fill
    ws_dashboard.cell(row=11, column=2).border = header_border
    ws_dashboard.merge_cells('B11:G11')
    
    ws_dashboard.cell(row=11, column=8, value="Adet").font = Font(bold=True, size=11, color='374151')
    ws_dashboard.cell(row=11, column=8).fill = header_fill
    ws_dashboard.cell(row=11, column=8).border = header_border
    ws_dashboard.cell(row=11, column=8).alignment = Alignment(horizontal='center')
    
    ws_dashboard.cell(row=11, column=9, value="Oran").font = Font(bold=True, size=11, color='374151')
    ws_dashboard.cell(row=11, column=9).fill = header_fill
    ws_dashboard.cell(row=11, column=9).border = header_border
    ws_dashboard.cell(row=11, column=9).alignment = Alignment(horizontal='center')
    ws_dashboard.merge_cells('I11:J11')
    
    # Kategori data
    row_colors = ['e0e7ff', 'ddd6fe', 'dcfce7', 'fef3c7', 'fee2e2', 'e0f2fe']
    for idx, (kategori, count) in enumerate(kategori_list[:10]):  # Top 10 categories
        row_num = 12 + idx
        percentage = round((count / total_count) * 100) if total_count > 0 else 0
        
        color = row_colors[idx % len(row_colors)]
        row_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        # Kategori name
        kat_cell = ws_dashboard.cell(row=row_num, column=2, value=kategori)
        kat_cell.font = Font(size=10, color='374151')
        kat_cell.fill = row_fill
        ws_dashboard.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=7)
        
        # Count
        count_cell = ws_dashboard.cell(row=row_num, column=8, value=count)
        count_cell.font = Font(size=10, color='374151', bold=True)
        count_cell.fill = row_fill
        count_cell.alignment = Alignment(horizontal='center')
        
        # Percentage
        pct_cell = ws_dashboard.cell(row=row_num, column=9, value=f"%{percentage}")
        pct_cell.font = Font(size=10, color='6b7280')
        pct_cell.fill = row_fill
        pct_cell.alignment = Alignment(horizontal='center')
        ws_dashboard.merge_cells(start_row=row_num, start_column=9, end_row=row_num, end_column=10)
    
    # Set wider columns for readability
    ws_dashboard.column_dimensions['B'].width = 8
    ws_dashboard.column_dimensions['C'].width = 8
    ws_dashboard.column_dimensions['D'].width = 8
    ws_dashboard.column_dimensions['E'].width = 8
    ws_dashboard.column_dimensions['H'].width = 8
    ws_dashboard.column_dimensions['I'].width = 8
    
    # ===== RAPORLAR (DATA) SHEET =====
    ws_data = wb.create_sheet("Raporlar")
    
    headers = [
        "Rapor No", "Ekipman Adı", "Kategori", "Firma", "Lokasyon",
        "Marka/Model", "Seri No", "Alt Kategori", "Periyot",
        "Geçerlilik Tarihi", "Uygunluk", "Proje", "Şehir", "Açıklama", "Oluşturma Tarihi"
    ]
    
    data_header_fill = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
    data_header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.fill = data_header_fill
        cell.font = data_header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, rapor in enumerate(raporlar, 2):
        ws_data.cell(row=row_idx, column=1, value=rapor.get("rapor_no", ""))
        ws_data.cell(row=row_idx, column=2, value=rapor.get("ekipman_adi", ""))
        ws_data.cell(row=row_idx, column=3, value=rapor.get("kategori", ""))
        ws_data.cell(row=row_idx, column=4, value=rapor.get("firma", ""))
        ws_data.cell(row=row_idx, column=5, value=rapor.get("lokasyon", ""))
        ws_data.cell(row=row_idx, column=6, value=rapor.get("marka_model", ""))
        ws_data.cell(row=row_idx, column=7, value=rapor.get("seri_no", ""))
        ws_data.cell(row=row_idx, column=8, value=rapor.get("alt_kategori", ""))
        ws_data.cell(row=row_idx, column=9, value=rapor.get("periyot", ""))
        ws_data.cell(row=row_idx, column=10, value=rapor.get("gecerlilik_tarihi", ""))
        ws_data.cell(row=row_idx, column=11, value=rapor.get("uygunluk", ""))
        ws_data.cell(row=row_idx, column=12, value=rapor.get("proje_adi", ""))
        ws_data.cell(row=row_idx, column=13, value=rapor.get("sehir", ""))
        ws_data.cell(row=row_idx, column=14, value=rapor.get("aciklama", ""))
        created_at = rapor.get("created_at", "")
        if isinstance(created_at, str):
            try:
                created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            except:
                created_at = None
        ws_data.cell(row=row_idx, column=15, value=created_at.strftime("%Y-%m-%d %H:%M") if created_at else "")
    
    # Auto-fit columns in data sheet
    for col in ws_data.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_data.column_dimensions[column].width = adjusted_width
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create filename with filter info
    filter_parts = []
    if request.proje_id and request.proje_id != 'all':
        filter_parts.append("proje")
    if request.sehir and request.sehir != 'all':
        filter_parts.append(request.sehir)
    if request.firma and request.firma != 'all':
        filter_parts.append("firma")
    
    filter_suffix = "_".join(filter_parts) if filter_parts else "tum"
    filename = f"dashboard_{filter_suffix}_{total_count}_adet.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export-all")
async def export_excel(current_user: dict = Depends(get_current_user)):
    """Tüm raporları Excel'e aktar"""
    raporlar = await db.raporlar.find({}, {"_id": 0}).to_list(10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Raporlar"
    
    headers = [
        "Rapor No", "Ekipman Adı", "Kategori", "Firma", "Lokasyon",
        "Marka/Model", "Seri No", "Alt Kategori", "Periyot",
        "Geçerlilik Tarihi", "Uygunluk", "Açıklama", "Oluşturma Tarihi"
    ]
    
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, rapor in enumerate(raporlar, 2):
        ws.cell(row=row_idx, column=1, value=rapor.get("rapor_no", ""))
        ws.cell(row=row_idx, column=2, value=rapor.get("ekipman_adi", ""))
        ws.cell(row=row_idx, column=3, value=rapor.get("kategori", ""))
        ws.cell(row=row_idx, column=4, value=rapor.get("firma", ""))
        ws.cell(row=row_idx, column=5, value=rapor.get("lokasyon", ""))
        ws.cell(row=row_idx, column=6, value=rapor.get("marka_model", ""))
        ws.cell(row=row_idx, column=7, value=rapor.get("seri_no", ""))
        ws.cell(row=row_idx, column=8, value=rapor.get("alt_kategori", ""))
        ws.cell(row=row_idx, column=9, value=rapor.get("periyot", ""))
        ws.cell(row=row_idx, column=10, value=rapor.get("gecerlilik_tarihi", ""))
        ws.cell(row=row_idx, column=11, value=rapor.get("uygunluk", ""))
        ws.cell(row=row_idx, column=12, value=rapor.get("aciklama", ""))
        created_at = rapor.get("created_at", "")
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        ws.cell(row=row_idx, column=13, value=created_at.strftime("%Y-%m-%d %H:%M") if created_at else "")
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=raporlar.xlsx"}
    )

@router.get("/template")
async def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapor Şablonu"
    
    headers = [
        "Şehir", "Ekipman Adı", "Kategori", "Firma", "Lokasyon", "Marka/Model",
        "Seri No", "Alt Kategori", "Periyot", "Geçerlilik Tarihi",
        "Uygunluk", "Açıklama"
    ]
    
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    examples = [
        ["İstanbul", "Asansör A1", "Asansör", "ABC Firma", "İstanbul Ofis", "Otis 2000",
         "SN12345", "Yolcu Asansörü", "6 Aylık", "2025-12-31", "Uygun", "Örnek açıklama"],
        ["Ankara", "Hidrofor H2", "Basınçlı Kaplar", "XYZ Ltd", "Ankara Fabrika", "Grundfos",
         "SN67890", "Hidrofor", "3 Aylık", "2025-09-30", "Uygun Değil", ""],
        ["İzmir", "Forklift F3", "Forklift", "Test Firma", "İzmir Depo", "Toyota", "", "", "", "", "", ""]
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col, value in enumerate(example, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    ws_cities = wb.create_sheet("Şehir Listesi")
    ws_cities.cell(row=1, column=1, value="Geçerli Şehirler (Büyük/küçük harf önemli değil)")
    ws_cities.cell(row=1, column=1).font = Font(bold=True, size=12)
    
    for idx, sehir in enumerate(SEHIRLER, 2):
        ws_cities.cell(row=idx, column=1, value=sehir["isim"])
    
    ws_cities.column_dimensions['A'].width = 25
    
    for col in ws.columns:
        column = col[0].column_letter
        ws.column_dimensions[column].width = 20
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rapor_sablonu.xlsx"}
    )

@router.post("/import")
async def import_excel(
    file: UploadFile = File(...),
    proje_id: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "inspector"]:
        raise HTTPException(status_code=403, detail="Excel içe aktarma yetkiniz yok")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sadece Excel dosyaları yüklenebilir")
    
    proje = await db.projeler.find_one({"id": proje_id}, {"_id": 0})
    if not proje:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    proje_adi = proje.get("proje_adi", "")
    
    content = await file.read()
    excel_file = io.BytesIO(content)
    
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        imported_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):
                continue
            
            try:
                def get_cell(index):
                    return row[index] if index < len(row) and row[index] is not None else None
                
                sehir = str(get_cell(0)) if get_cell(0) else None
                ekipman_adi = str(get_cell(1)) if get_cell(1) else ""
                kategori = str(get_cell(2)) if get_cell(2) else ""
                firma = str(get_cell(3)) if get_cell(3) else ""
                lokasyon = str(get_cell(4)) if get_cell(4) else None
                marka_model = str(get_cell(5)) if get_cell(5) else None
                seri_no = str(get_cell(6)) if get_cell(6) else None
                alt_kategori = str(get_cell(7)) if get_cell(7) else None
                periyot = str(get_cell(8)) if get_cell(8) else None
                gecerlilik_tarihi = str(get_cell(9)) if get_cell(9) else None
                uygunluk = str(get_cell(10)) if get_cell(10) else None
                aciklama = str(get_cell(11)) if get_cell(11) else None
                
                if not ekipman_adi or not kategori or not firma:
                    errors.append(f"Satır {row_idx}: Zorunlu alanlar eksik (Ekipman Adı, Kategori, Firma)")
                    continue
                
                if not sehir:
                    errors.append(f"Satır {row_idx}: Şehir alanı zorunludur")
                    continue
                
                def turkish_lower(text):
                    text = text.strip()
                    text = text.replace('İ', 'i').replace('I', 'ı')
                    return text.lower()
                
                def normalize_turkish(text):
                    turkish_map = str.maketrans('ıİiIğĞüÜşŞöÖçÇ', 'iiiigguussoocc')
                    return turkish_lower(text).translate(turkish_map)
                
                sehir_normalized = normalize_turkish(sehir)
                
                sehir_obj = None
                for s in SEHIRLER:
                    if normalize_turkish(s["isim"]) == sehir_normalized:
                        sehir_obj = s
                        sehir = s["isim"]
                        break
                
                if not sehir_obj:
                    errors.append(f"Satır {row_idx}: Geçersiz şehir - '{sehir}'")
                    continue
                
                rapor_no = await generate_rapor_no(sehir)
                
                rapor_data = {
                    "id": str(uuid.uuid4()),
                    "rapor_no": rapor_no,
                    "proje_id": proje_id,
                    "proje_adi": proje_adi,
                    "sehir": sehir,
                    "sehir_kodu": sehir_obj["kod"],
                    "ekipman_adi": ekipman_adi,
                    "kategori": kategori,
                    "alt_kategori": alt_kategori,
                    "firma": firma,
                    "lokasyon": lokasyon,
                    "marka_model": marka_model,
                    "seri_no": seri_no,
                    "periyot": periyot,
                    "gecerlilik_tarihi": gecerlilik_tarihi,
                    "uygunluk": uygunluk,
                    "aciklama": aciklama,
                    "durum": "Aktif",
                    "created_by": current_user["id"],
                    "created_by_username": current_user.get("username", current_user.get("email", "")),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.raporlar.insert_one(rapor_data)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Satır {row_idx}: {str(e)}")
        
        return {
            "message": f"{imported_count} rapor başarıyla içe aktarıldı",
            "imported_count": imported_count,
            "errors": errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel işleme hatası: {str(e)}")
