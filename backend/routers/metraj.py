"""
Metraj Cetveli Router - Bill of Quantities (BOQ) Module
CRUD operations and Excel export for metraj data
"""

from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone
from bson import ObjectId
import io
import uuid

from database import db
from routers.auth import get_current_user

router = APIRouter(prefix="/metraj", tags=["Metraj Cetveli"])


# ==================== MODELS ====================

class MetrajSatiri(BaseModel):
    """Single row in metraj table"""
    id: Optional[str] = None
    poz_no: str = ""
    malzeme_adi: str = ""
    birim: str = "Adet"  # Adet, m, m², m³, kg, ton, lt, takım, vb.
    miktar: float = 0.0
    birim_fiyat: float = 0.0
    birim_agirlik: Optional[float] = None  # kg/m, kg/m² vb.
    toplam: float = 0.0
    toplam_agirlik: Optional[float] = None
    aciklama: str = ""
    sira_no: int = 0


class MetrajCetveli(BaseModel):
    """Full metraj document"""
    id: Optional[str] = None
    proje_id: Optional[str] = None
    rapor_id: Optional[str] = None
    baslik: str = "Metraj Cetveli"
    aciklama: str = ""
    satirlar: List[MetrajSatiri] = []
    genel_toplam: float = 0.0
    genel_agirlik: Optional[float] = None
    created_at: Optional[str] = None
    updated_at: Optional[str] = None
    created_by: Optional[str] = None


class MetrajSatiriCreate(BaseModel):
    """Create/Update satir request"""
    poz_no: str = ""
    malzeme_adi: str = ""
    birim: str = "Adet"
    miktar: float = 0.0
    birim_fiyat: float = 0.0
    birim_agirlik: Optional[float] = None
    aciklama: str = ""


class MetrajCetveliCreate(BaseModel):
    """Create metraj request"""
    proje_id: Optional[str] = None
    rapor_id: Optional[str] = None
    baslik: str = "Metraj Cetveli"
    aciklama: str = ""


class MetrajSatiriUpdate(BaseModel):
    """Update single satir"""
    poz_no: Optional[str] = None
    malzeme_adi: Optional[str] = None
    birim: Optional[str] = None
    miktar: Optional[float] = None
    birim_fiyat: Optional[float] = None
    birim_agirlik: Optional[float] = None
    aciklama: Optional[str] = None


# ==================== HELPER FUNCTIONS ====================

def calculate_satir_totals(satir: dict) -> dict:
    """Calculate totals for a single row"""
    miktar = satir.get("miktar", 0) or 0
    birim_fiyat = satir.get("birim_fiyat", 0) or 0
    birim_agirlik = satir.get("birim_agirlik")
    
    satir["toplam"] = round(miktar * birim_fiyat, 2)
    
    if birim_agirlik is not None:
        satir["toplam_agirlik"] = round(miktar * birim_agirlik, 2)
    else:
        satir["toplam_agirlik"] = None
    
    return satir


def calculate_cetvel_totals(satirlar: List[dict]) -> tuple:
    """Calculate grand totals for entire table"""
    genel_toplam = sum(s.get("toplam", 0) or 0 for s in satirlar)
    
    agirliklar = [s.get("toplam_agirlik") for s in satirlar if s.get("toplam_agirlik") is not None]
    genel_agirlik = sum(agirliklar) if agirliklar else None
    
    return round(genel_toplam, 2), round(genel_agirlik, 2) if genel_agirlik else None


# ==================== CRUD ENDPOINTS ====================

@router.post("/", response_model=dict)
async def create_metraj_cetveli(
    data: MetrajCetveliCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new metraj cetveli"""
    
    cetvel_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    cetvel = {
        "id": cetvel_id,
        "proje_id": data.proje_id,
        "rapor_id": data.rapor_id,
        "baslik": data.baslik,
        "aciklama": data.aciklama,
        "satirlar": [],
        "genel_toplam": 0.0,
        "genel_agirlik": None,
        "created_at": now,
        "updated_at": now,
        "created_by": current_user.get("id")
    }
    
    await db.metraj_cetvelleri.insert_one(cetvel)
    
    return {"id": cetvel_id, "message": "Metraj cetveli oluşturuldu"}


@router.get("/", response_model=List[dict])
async def list_metraj_cetvelleri(
    proje_id: Optional[str] = None,
    rapor_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """List all metraj cetvelleri"""
    
    query = {}
    if proje_id:
        query["proje_id"] = proje_id
    if rapor_id:
        query["rapor_id"] = rapor_id
    
    cetveller = await db.metraj_cetvelleri.find(query, {"_id": 0}).to_list(None)
    
    return cetveller


@router.get("/{cetvel_id}", response_model=dict)
async def get_metraj_cetveli(
    cetvel_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get a specific metraj cetveli"""
    
    cetvel = await db.metraj_cetvelleri.find_one({"id": cetvel_id}, {"_id": 0})
    
    if not cetvel:
        raise HTTPException(status_code=404, detail="Metraj cetveli bulunamadı")
    
    return cetvel


@router.put("/{cetvel_id}", response_model=dict)
async def update_metraj_cetveli(
    cetvel_id: str,
    data: MetrajCetveliCreate,
    current_user: dict = Depends(get_current_user)
):
    """Update metraj cetveli metadata"""
    
    cetvel = await db.metraj_cetvelleri.find_one({"id": cetvel_id})
    if not cetvel:
        raise HTTPException(status_code=404, detail="Metraj cetveli bulunamadı")
    
    update_data = {
        "baslik": data.baslik,
        "aciklama": data.aciklama,
        "proje_id": data.proje_id,
        "rapor_id": data.rapor_id,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.metraj_cetvelleri.update_one({"id": cetvel_id}, {"$set": update_data})
    
    return {"message": "Metraj cetveli güncellendi"}


@router.delete("/{cetvel_id}")
async def delete_metraj_cetveli(
    cetvel_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a metraj cetveli"""
    
    result = await db.metraj_cetvelleri.delete_one({"id": cetvel_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Metraj cetveli bulunamadı")
    
    return {"message": "Metraj cetveli silindi"}


# ==================== SATIR (ROW) ENDPOINTS ====================

@router.post("/{cetvel_id}/satir", response_model=dict)
async def add_satir(
    cetvel_id: str,
    data: MetrajSatiriCreate,
    current_user: dict = Depends(get_current_user)
):
    """Add a new row to metraj cetveli"""
    
    cetvel = await db.metraj_cetvelleri.find_one({"id": cetvel_id})
    if not cetvel:
        raise HTTPException(status_code=404, detail="Metraj cetveli bulunamadı")
    
    satirlar = cetvel.get("satirlar", [])
    
    # Generate new row
    satir_id = str(uuid.uuid4())
    sira_no = len(satirlar) + 1
    
    satir = {
        "id": satir_id,
        "poz_no": data.poz_no,
        "malzeme_adi": data.malzeme_adi,
        "birim": data.birim,
        "miktar": data.miktar,
        "birim_fiyat": data.birim_fiyat,
        "birim_agirlik": data.birim_agirlik,
        "aciklama": data.aciklama,
        "sira_no": sira_no
    }
    
    # Calculate totals
    satir = calculate_satir_totals(satir)
    
    satirlar.append(satir)
    
    # Recalculate grand totals
    genel_toplam, genel_agirlik = calculate_cetvel_totals(satirlar)
    
    await db.metraj_cetvelleri.update_one(
        {"id": cetvel_id},
        {
            "$set": {
                "satirlar": satirlar,
                "genel_toplam": genel_toplam,
                "genel_agirlik": genel_agirlik,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    return {"id": satir_id, "satir": satir, "message": "Satır eklendi"}


@router.put("/{cetvel_id}/satir/{satir_id}", response_model=dict)
async def update_satir(
    cetvel_id: str,
    satir_id: str,
    data: MetrajSatiriUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a specific row"""
    
    cetvel = await db.metraj_cetvelleri.find_one({"id": cetvel_id})
    if not cetvel:
        raise HTTPException(status_code=404, detail="Metraj cetveli bulunamadı")
    
    satirlar = cetvel.get("satirlar", [])
    satir_index = next((i for i, s in enumerate(satirlar) if s.get("id") == satir_id), None)
    
    if satir_index is None:
        raise HTTPException(status_code=404, detail="Satır bulunamadı")
    
    # Update fields
    satir = satirlar[satir_index]
    update_dict = data.dict(exclude_unset=True)
    
    for key, value in update_dict.items():
        if value is not None:
            satir[key] = value
    
    # Recalculate row totals
    satir = calculate_satir_totals(satir)
    satirlar[satir_index] = satir
    
    # Recalculate grand totals
    genel_toplam, genel_agirlik = calculate_cetvel_totals(satirlar)
    
    await db.metraj_cetvelleri.update_one(
        {"id": cetvel_id},
        {
            "$set": {
                "satirlar": satirlar,
                "genel_toplam": genel_toplam,
                "genel_agirlik": genel_agirlik,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    return {"satir": satir, "genel_toplam": genel_toplam, "message": "Satır güncellendi"}


@router.delete("/{cetvel_id}/satir/{satir_id}")
async def delete_satir(
    cetvel_id: str,
    satir_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a specific row"""
    
    cetvel = await db.metraj_cetvelleri.find_one({"id": cetvel_id})
    if not cetvel:
        raise HTTPException(status_code=404, detail="Metraj cetveli bulunamadı")
    
    satirlar = cetvel.get("satirlar", [])
    satirlar = [s for s in satirlar if s.get("id") != satir_id]
    
    # Renumber rows
    for i, satir in enumerate(satirlar):
        satir["sira_no"] = i + 1
    
    # Recalculate grand totals
    genel_toplam, genel_agirlik = calculate_cetvel_totals(satirlar)
    
    await db.metraj_cetvelleri.update_one(
        {"id": cetvel_id},
        {
            "$set": {
                "satirlar": satirlar,
                "genel_toplam": genel_toplam,
                "genel_agirlik": genel_agirlik,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    return {"message": "Satır silindi", "genel_toplam": genel_toplam}


@router.post("/{cetvel_id}/satir/{satir_id}/duplicate", response_model=dict)
async def duplicate_satir(
    cetvel_id: str,
    satir_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Duplicate a specific row"""
    
    cetvel = await db.metraj_cetvelleri.find_one({"id": cetvel_id})
    if not cetvel:
        raise HTTPException(status_code=404, detail="Metraj cetveli bulunamadı")
    
    satirlar = cetvel.get("satirlar", [])
    original = next((s for s in satirlar if s.get("id") == satir_id), None)
    
    if not original:
        raise HTTPException(status_code=404, detail="Satır bulunamadı")
    
    # Create duplicate
    new_satir = original.copy()
    new_satir["id"] = str(uuid.uuid4())
    new_satir["sira_no"] = len(satirlar) + 1
    new_satir["poz_no"] = f"{original.get('poz_no', '')}-KOPYA"
    
    satirlar.append(new_satir)
    
    # Recalculate grand totals
    genel_toplam, genel_agirlik = calculate_cetvel_totals(satirlar)
    
    await db.metraj_cetvelleri.update_one(
        {"id": cetvel_id},
        {
            "$set": {
                "satirlar": satirlar,
                "genel_toplam": genel_toplam,
                "genel_agirlik": genel_agirlik,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    return {"id": new_satir["id"], "satir": new_satir, "message": "Satır kopyalandı"}


@router.put("/{cetvel_id}/bulk-update", response_model=dict)
async def bulk_update_satirlar(
    cetvel_id: str,
    satirlar: List[dict],
    current_user: dict = Depends(get_current_user)
):
    """Bulk update all rows at once (for frontend table sync)"""
    
    cetvel = await db.metraj_cetvelleri.find_one({"id": cetvel_id})
    if not cetvel:
        raise HTTPException(status_code=404, detail="Metraj cetveli bulunamadı")
    
    # Process all rows
    processed_satirlar = []
    for i, satir in enumerate(satirlar):
        if not satir.get("id"):
            satir["id"] = str(uuid.uuid4())
        satir["sira_no"] = i + 1
        satir = calculate_satir_totals(satir)
        processed_satirlar.append(satir)
    
    # Calculate grand totals
    genel_toplam, genel_agirlik = calculate_cetvel_totals(processed_satirlar)
    
    await db.metraj_cetvelleri.update_one(
        {"id": cetvel_id},
        {
            "$set": {
                "satirlar": processed_satirlar,
                "genel_toplam": genel_toplam,
                "genel_agirlik": genel_agirlik,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    return {
        "message": f"{len(processed_satirlar)} satır güncellendi",
        "genel_toplam": genel_toplam,
        "genel_agirlik": genel_agirlik
    }


# ==================== EXCEL EXPORT ====================

@router.get("/{cetvel_id}/export-excel")
async def export_metraj_excel(
    cetvel_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Export metraj cetveli as professionally formatted Excel file"""
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    cetvel = await db.metraj_cetvelleri.find_one({"id": cetvel_id}, {"_id": 0})
    if not cetvel:
        raise HTTPException(status_code=404, detail="Metraj cetveli bulunamadı")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Metraj Cetveli"
    
    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(name="Arial", size=11, bold=True)
    cell_font = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = cetvel.get("baslik", "Metraj Cetveli")
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    # Description
    if cetvel.get("aciklama"):
        ws.merge_cells('A2:I2')
        ws['A2'] = cetvel.get("aciklama")
        ws['A2'].alignment = center_align
    
    # Headers - Row 4
    headers = [
        ("Sıra", 8),
        ("Poz No", 12),
        ("Malzeme Adı", 35),
        ("Birim", 10),
        ("Miktar", 12),
        ("Birim Fiyat (₺)", 15),
        ("Birim Ağırlık", 14),
        ("Toplam (₺)", 15),
        ("Açıklama", 25)
    ]
    
    header_row = 4
    for col, (header_text, width) in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Data rows
    satirlar = cetvel.get("satirlar", [])
    data_start_row = header_row + 1
    
    for row_idx, satir in enumerate(satirlar, data_start_row):
        row_data = [
            satir.get("sira_no", row_idx - data_start_row + 1),
            satir.get("poz_no", ""),
            satir.get("malzeme_adi", ""),
            satir.get("birim", ""),
            satir.get("miktar", 0),
            satir.get("birim_fiyat", 0),
            satir.get("birim_agirlik", "") if satir.get("birim_agirlik") else "-",
            satir.get("toplam", 0),
            satir.get("aciklama", "")
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = cell_font
            cell.border = thin_border
            
            # Alignment
            if col in [1, 4]:  # Sıra, Birim
                cell.alignment = center_align
            elif col in [5, 6, 7, 8]:  # Numbers
                cell.alignment = right_align
                if isinstance(value, (int, float)) and col in [6, 8]:
                    cell.number_format = '#,##0.00 ₺'
                elif isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = left_align
    
    # Totals row
    total_row = data_start_row + len(satirlar)
    
    ws.merge_cells(f'A{total_row}:G{total_row}')
    ws[f'A{total_row}'] = "GENEL TOPLAM"
    ws[f'A{total_row}'].font = total_font
    ws[f'A{total_row}'].fill = total_fill
    ws[f'A{total_row}'].alignment = right_align
    ws[f'A{total_row}'].border = thin_border
    
    # Fill merged cells border
    for col in range(2, 8):
        ws.cell(row=total_row, column=col).fill = total_fill
        ws.cell(row=total_row, column=col).border = thin_border
    
    # Total value
    ws[f'H{total_row}'] = cetvel.get("genel_toplam", 0)
    ws[f'H{total_row}'].font = total_font
    ws[f'H{total_row}'].fill = total_fill
    ws[f'H{total_row}'].alignment = right_align
    ws[f'H{total_row}'].border = thin_border
    ws[f'H{total_row}'].number_format = '#,##0.00 ₺'
    
    ws[f'I{total_row}'] = ""
    ws[f'I{total_row}'].fill = total_fill
    ws[f'I{total_row}'].border = thin_border
    
    # Weight totals if applicable
    if cetvel.get("genel_agirlik"):
        weight_row = total_row + 1
        ws.merge_cells(f'A{weight_row}:G{weight_row}')
        ws[f'A{weight_row}'] = "TOPLAM AĞIRLIK (kg)"
        ws[f'A{weight_row}'].font = total_font
        ws[f'A{weight_row}'].fill = total_fill
        ws[f'A{weight_row}'].alignment = right_align
        
        ws[f'H{weight_row}'] = cetvel.get("genel_agirlik", 0)
        ws[f'H{weight_row}'].font = total_font
        ws[f'H{weight_row}'].fill = total_fill
        ws[f'H{weight_row}'].alignment = right_align
        ws[f'H{weight_row}'].number_format = '#,##0.00'
    
    # Footer with date
    footer_row = total_row + 3
    ws[f'A{footer_row}'] = f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws[f'A{footer_row}'].font = Font(name="Arial", size=9, italic=True)
    
    # Freeze header row
    ws.freeze_panes = f'A{data_start_row}'
    
    # Print settings
    ws.print_title_rows = f'{header_row}:{header_row}'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"metraj_cetveli_{cetvel.get('baslik', 'export').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== BIRIM OPTIONS ====================

@router.get("/birimler/liste")
async def get_birim_options():
    """Get list of available unit types"""
    return {
        "birimler": [
            {"value": "Adet", "label": "Adet"},
            {"value": "m", "label": "Metre (m)"},
            {"value": "m²", "label": "Metrekare (m²)"},
            {"value": "m³", "label": "Metreküp (m³)"},
            {"value": "kg", "label": "Kilogram (kg)"},
            {"value": "ton", "label": "Ton"},
            {"value": "lt", "label": "Litre (lt)"},
            {"value": "takım", "label": "Takım"},
            {"value": "set", "label": "Set"},
            {"value": "paket", "label": "Paket"},
            {"value": "kutu", "label": "Kutu"},
            {"value": "top", "label": "Top"},
            {"value": "rulo", "label": "Rulo"},
        ]
    }
