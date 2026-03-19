"""
Kombinasyonlar Router - Şans Topu Tahmin Üretici
Rastgele kombinasyon üretme, arama ve Excel'e kaydetme işlemleri

Şans Topu Kuralları:
- 5 ana sayı: 1-34 arası (birbirinden farklı)
- 1 bonus sayı: 1-14 arası (rastgele)
- Toplam benzersiz beşli kombinasyon sayısı: C(34,5) = 278.256
- Her beşli kombinasyona rastgele bonus ile: ~1.030.144 kombinasyon
"""

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import List, Optional, Dict, Any, Set, Tuple
import random
from itertools import combinations
import io
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/kombinasyonlar", tags=["Kombinasyonlar"])

# Global cache for combinations
COMBINATIONS_CACHE: Dict[int, List[int]] = {}
# Track unique 5-number combinations
UNIQUE_FIVES: Set[Tuple[int, ...]] = set()

# Default target: 1,030,144 unique combinations
DEFAULT_COMBINATION_COUNT = 1030144


class GenerateRequest(BaseModel):
    num_combinations: int = DEFAULT_COMBINATION_COUNT
    seed: Optional[int] = None


class SearchRequest(BaseModel):
    combination_str: str


class FilterRequest(BaseModel):
    main_numbers: List[int]
    bonus_number: Optional[int] = None


class GenerateResponse(BaseModel):
    generated_count: int
    total_count: int
    message: str


class SearchResult(BaseModel):
    found: bool
    indices: List[int]
    combination: Optional[List[int]]
    message: str


class CombinationItem(BaseModel):
    index: int
    main_numbers: List[int]
    bonus_number: int
    formatted: str


class StatsResponse(BaseModel):
    total_combinations: int
    unique_five_combos: Optional[int] = 0
    default_target: Optional[int] = DEFAULT_COMBINATION_COUNT


@router.get("/stats", response_model=StatsResponse)
async def get_stats():
    """Get current combination statistics"""
    return {
        "total_combinations": len(COMBINATIONS_CACHE),
        "unique_five_combos": len(UNIQUE_FIVES),
        "default_target": DEFAULT_COMBINATION_COUNT
    }


@router.post("/generate", response_model=GenerateResponse)
async def generate_combinations(request: GenerateRequest):
    """
    Generate new combinations with UNIQUE 5-number sets.
    Each 5-number combination appears only once, bonus is random.
    Target: 1,030,144 unique combinations.
    """
    global COMBINATIONS_CACHE, UNIQUE_FIVES
    
    if request.seed:
        random.seed(request.seed)
    
    num = request.num_combinations
    if num <= 0:
        raise HTTPException(status_code=400, detail="Kombinasyon sayısı pozitif olmalıdır")
    
    # C(34,5) = 278,256 unique 5-number combinations possible
    max_unique_fives = 278256
    
    if num > max_unique_fives * 14:  # Each five can have 14 different bonuses
        raise HTTPException(
            status_code=400, 
            detail=f"Maksimum {max_unique_fives * 14:,} kombinasyon üretilebilir (278.256 beşli × 14 bonus)"
        )
    
    # Clear existing if regenerating
    COMBINATIONS_CACHE.clear()
    UNIQUE_FIVES.clear()
    
    generated = 0
    attempts = 0
    max_attempts = num * 3  # Prevent infinite loop
    
    while generated < num and attempts < max_attempts:
        attempts += 1
        
        # Generate unique 5-number combination
        main_numbers = tuple(sorted(random.sample(range(1, 35), 5)))
        
        # Check if this 5-number combo already exists
        if main_numbers in UNIQUE_FIVES:
            continue
        
        # Add to unique set
        UNIQUE_FIVES.add(main_numbers)
        
        # Random bonus (1-14)
        bonus_number = random.randint(1, 14)
        
        # Store combination
        generated += 1
        combination = list(main_numbers) + [bonus_number]
        COMBINATIONS_CACHE[generated] = combination
    
    return {
        "generated_count": generated,
        "total_count": len(COMBINATIONS_CACHE),
        "message": f"{generated:,} benzersiz kombinasyon oluşturuldu. Her beşli farklı!"
    }


@router.post("/clear")
async def clear_cache():
    """Clear all combinations from cache"""
    global COMBINATIONS_CACHE, UNIQUE_FIVES
    count = len(COMBINATIONS_CACHE)
    COMBINATIONS_CACHE.clear()
    UNIQUE_FIVES.clear()
    return {
        "cleared_count": count,
        "message": f"{count:,} kombinasyon silindi. Önbellek temizlendi."
    }


@router.post("/search", response_model=SearchResult)
async def search_combination(request: SearchRequest):
    """Search for a combination in cache"""
    global COMBINATIONS_CACHE
    
    if not COMBINATIONS_CACHE:
        return {
            "found": False,
            "indices": [],
            "combination": None,
            "message": "Cache boş! Önce kombinasyon oluşturun."
        }
    
    combination_str = request.combination_str.strip()
    
    try:
        # Parse different formats
        if '+' in combination_str:
            # Format: 5,12,23,27,34+8
            main_part, bonus_part = combination_str.split('+')
            main_numbers = [int(num.strip()) for num in main_part.split(',')]
            bonus_number = int(bonus_part.strip())
        elif '-' in combination_str and combination_str.count('-') >= 4:
            # Format: 5-12-23-27-34-8
            numbers = [int(num.strip()) for num in combination_str.split('-')]
            main_numbers = numbers[:5]
            bonus_number = numbers[5]
        else:
            # Format: 5 12 23 27 34 8 or 5,12,23,27,34,8
            numbers = [int(num.strip()) for num in combination_str.replace(',', ' ').split()]
            if len(numbers) != 6:
                return {
                    "found": False,
                    "indices": [],
                    "combination": None,
                    "message": "Geçersiz format! 6 sayı girmelisiniz (5 ana + 1 bonus)."
                }
            main_numbers = numbers[:5]
            bonus_number = numbers[5]
        
        # Validate numbers
        if len(main_numbers) != 5:
            return {
                "found": False,
                "indices": [],
                "combination": None,
                "message": "5 ana sayı girilmelidir."
            }
        
        if any(num < 1 or num > 34 for num in main_numbers):
            return {
                "found": False,
                "indices": [],
                "combination": None,
                "message": "Ana sayılar 1-34 arasında olmalıdır."
            }
        
        if len(set(main_numbers)) != 5:
            return {
                "found": False,
                "indices": [],
                "combination": None,
                "message": "Ana sayılar birbirinden farklı olmalıdır."
            }
        
        if bonus_number < 1 or bonus_number > 14:
            return {
                "found": False,
                "indices": [],
                "combination": None,
                "message": "Bonus sayı 1-14 arasında olmalıdır."
            }
        
        # Sort and create target
        main_numbers.sort()
        target_combination = main_numbers + [bonus_number]
        
        # Search in cache
        found_indices = []
        for idx, comb in COMBINATIONS_CACHE.items():
            if comb == target_combination:
                found_indices.append(idx)
        
        if found_indices:
            return {
                "found": True,
                "indices": found_indices[:100],  # Return max 100 results
                "combination": target_combination,
                "message": f"Kombinasyon {len(found_indices)} kez bulundu!"
            }
        else:
            return {
                "found": False,
                "indices": [],
                "combination": target_combination,
                "message": "Kombinasyon bulunamadı."
            }
            
    except ValueError:
        return {
            "found": False,
            "indices": [],
            "combination": None,
            "message": "Geçersiz sayı formatı!"
        }
    except Exception as e:
        return {
            "found": False,
            "indices": [],
            "combination": None,
            "message": f"Arama hatası: {str(e)}"
        }


class FilterResult(BaseModel):
    found: bool
    total_found: int
    results: List[CombinationItem]
    message: str


@router.post("/filter", response_model=FilterResult)
async def filter_combinations(request: FilterRequest):
    """Filter combinations by main numbers and optional bonus"""
    global COMBINATIONS_CACHE
    
    if not COMBINATIONS_CACHE:
        return {
            "found": False,
            "total_found": 0,
            "results": [],
            "message": "Cache boş! Önce kombinasyon oluşturun."
        }
    
    main_numbers = request.main_numbers
    bonus_number = request.bonus_number
    
    # Validate main numbers
    if len(main_numbers) != 5:
        return {
            "found": False,
            "total_found": 0,
            "results": [],
            "message": "Tam olarak 5 ana sayı seçmelisiniz."
        }
    
    if any(num < 1 or num > 34 for num in main_numbers):
        return {
            "found": False,
            "total_found": 0,
            "results": [],
            "message": "Ana sayılar 1-34 arasında olmalıdır."
        }
    
    if len(set(main_numbers)) != 5:
        return {
            "found": False,
            "total_found": 0,
            "results": [],
            "message": "Ana sayılar birbirinden farklı olmalıdır."
        }
    
    # Validate bonus if provided
    if bonus_number is not None and (bonus_number < 1 or bonus_number > 14):
        return {
            "found": False,
            "total_found": 0,
            "results": [],
            "message": "Bonus sayı 1-14 arasında olmalıdır."
        }
    
    # Sort main numbers for comparison
    sorted_main = sorted(main_numbers)
    
    # Search in cache
    found_results = []
    for idx, comb in COMBINATIONS_CACHE.items():
        comb_main = comb[:5]
        comb_bonus = comb[5]
        
        # Check if main numbers match
        if comb_main == sorted_main:
            # If bonus specified, check it too
            if bonus_number is None or comb_bonus == bonus_number:
                found_results.append({
                    "index": idx,
                    "main_numbers": comb_main,
                    "bonus_number": comb_bonus,
                    "formatted": f"{comb_main[0]},{comb_main[1]},{comb_main[2]},{comb_main[3]},{comb_main[4]}+{comb_bonus}"
                })
    
    # Limit results to 100
    limited_results = found_results[:100]
    
    if found_results:
        if bonus_number is None:
            message = f"Seçilen 5 ana sayı ile {len(found_results)} kombinasyon bulundu (tüm bonus değerleri)"
        else:
            message = f"Kombinasyon {len(found_results)} kez bulundu!"
        
        return {
            "found": True,
            "total_found": len(found_results),
            "results": limited_results,
            "message": message
        }
    else:
        if bonus_number is None:
            message = "Bu ana sayılarla kombinasyon bulunamadı."
        else:
            message = "Bu kombinasyon bulunamadı."
        
        return {
            "found": False,
            "total_found": 0,
            "results": [],
            "message": message
        }


@router.get("/sample", response_model=List[CombinationItem])
async def get_sample_combinations(count: int = 5):
    """Get random sample of combinations"""
    global COMBINATIONS_CACHE
    
    if not COMBINATIONS_CACHE:
        return []
    
    sample_count = min(count, len(COMBINATIONS_CACHE))
    sample_indices = random.sample(list(COMBINATIONS_CACHE.keys()), sample_count)
    
    results = []
    for idx in sorted(sample_indices):
        comb = COMBINATIONS_CACHE[idx]
        results.append({
            "index": idx,
            "main_numbers": comb[:5],
            "bonus_number": comb[5],
            "formatted": f"{comb[0]},{comb[1]},{comb[2]},{comb[3]},{comb[4]}+{comb[5]}"
        })
    
    return results


@router.get("/combination/{index}", response_model=CombinationItem)
async def get_combination(index: int):
    """Get a specific combination by index"""
    global COMBINATIONS_CACHE
    
    if index not in COMBINATIONS_CACHE:
        raise HTTPException(
            status_code=404, 
            detail=f"{index}. kombinasyon bulunamadı! Lütfen 1-{len(COMBINATIONS_CACHE)} arasında bir sayı girin."
        )
    
    comb = COMBINATIONS_CACHE[index]
    return {
        "index": index,
        "main_numbers": comb[:5],
        "bonus_number": comb[5],
        "formatted": f"{comb[0]},{comb[1]},{comb[2]},{comb[3]},{comb[4]}+{comb[5]}"
    }


@router.get("/export-excel")
async def export_to_excel():
    """Export all combinations to Excel file"""
    global COMBINATIONS_CACHE
    
    if not COMBINATIONS_CACHE:
        raise HTTPException(status_code=400, detail="Önce kombinasyon oluşturmanız gerekiyor!")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Kombinasyonlar"
    
    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Sıra", "Sayı 1", "Sayı 2", "Sayı 3", "Sayı 4", "Sayı 5", "Bonus", "Format"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Data - limit to first 100000 for performance
    max_rows = min(100000, len(COMBINATIONS_CACHE))
    sorted_keys = sorted(COMBINATIONS_CACHE.keys())[:max_rows]
    
    for row_idx, key in enumerate(sorted_keys, 2):
        comb = COMBINATIONS_CACHE[key]
        row_data = [
            key,
            comb[0], comb[1], comb[2], comb[3], comb[4],
            comb[5],
            f"{comb[0]},{comb[1]},{comb[2]},{comb[3]},{comb[4]}+{comb[5]}"
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border
    
    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions['H'].width = 20  # Format column wider
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"sans_topu_kombinasyonlari_{len(COMBINATIONS_CACHE)}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
